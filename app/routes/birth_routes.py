import os
import io
import re
import uuid
import shutil
import datetime
from fastapi import APIRouter, Request, Depends, HTTPException, Query, Form, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc, func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import User, Village, PollingStation, Voter, BirthCertificate
from app.auth import get_current_user_optional, get_current_user, require_admin, require_admin_or_officer
from app.audit import log_activity
from app.timezone_utils import get_cambodia_now, get_cambodia_today, get_cambodia_today_str

router = APIRouter()

templates_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
templates = Jinja2Templates(directory=templates_path)

UPLOAD_DIR_BIRTH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static", "uploads", "birth_certificates")
os.makedirs(UPLOAD_DIR_BIRTH, exist_ok=True)

def save_birth_attachment(upload_file: UploadFile) -> str:
    if not upload_file or not upload_file.filename:
        return None
    ext = os.path.splitext(upload_file.filename)[1].lower()
    allowed_exts = [".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"]
    if ext not in allowed_exts:
        ext = ".pdf" if "pdf" in (upload_file.content_type or "").lower() else ".jpg"
    filename = f"birth_{uuid.uuid4().hex[:14]}{ext}"
    file_path = os.path.join(UPLOAD_DIR_BIRTH, filename)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(upload_file.file, f)
    return f"/static/uploads/birth_certificates/{filename}"

def sanitize_id_backend(raw_id: str) -> str:
    if not raw_id:
        raise HTTPException(status_code=400, detail="សូមបញ្ចូលលេខអត្តសញ្ញាណប័ណ្ណ ឬឯកសារបញ្ជាក់អត្តសញ្ញាណ")
    khmer_map = str.maketrans("០១២៣៤៥៦៧៨៩", "0123456789")
    converted = raw_id.strip().translate(khmer_map)
    clean = re.sub(r"[^0-9]", "", converted)
    if not clean or len(clean) not in [7, 9]:
        raise HTTPException(
            status_code=400,
            detail=f"លេខអត្តសញ្ញាណប័ណ្ណត្រូវតែមាន ៩ ខ្ទង់ (ឬ ៧ ខ្ទង់សម្រាប់ ឯ.អ)"
        )
    return clean

KHMER_MONTH_NAMES = {
    1: "មករា", 2: "កុម្ភៈ", 3: "មីនា", 4: "មេសា",
    5: "ឧសភា", 6: "មិថុនា", 7: "កក្កដា", 8: "សីហា",
    9: "កញ្ញា", 10: "តុលា", 11: "វិច្ឆិកា", 12: "ធ្នូ"
}

def parse_created_dt(b):
    if not b:
        return None
    # Check registered_date first if available
    reg_d = getattr(b, "registered_date", None)
    if reg_d and str(reg_d).strip():
        parts = [int(p) for p in str(reg_d).strip().split("-") if p.isdigit()]
        if len(parts) >= 3:
            return datetime.datetime(parts[0], parts[1], parts[2])
        elif len(parts) == 2:
            return datetime.datetime(parts[0], parts[1], 1)
        elif len(parts) == 1:
            return datetime.datetime(parts[0], 1, 1)

    if not b.created_at:
        return None
    if isinstance(b.created_at, datetime.datetime):
        return b.created_at
    if isinstance(b.created_at, str):
        try:
            return datetime.datetime.fromisoformat(b.created_at.replace("Z", "+00:00").split(".")[0])
        except Exception:
            return None
    return None

@router.get("/birth-certificates", response_class=HTMLResponse)
def birth_certificates_list(
    request: Request,
    q: str = Query("", description="Search by name, certificate no, parent"),
    village_id: str = Query("", description="Village filter"),
    eligibility: str = Query("all", description="all, turning_18_this_year, turning_18_next_year, eligible_now, under_18"),
    reg_status: str = Query("all", description="all, registered, unregistered"),
    period: str = Query("all", description="all, this_month, last_month, this_year, last_year"),
    reg_year: str = Query("", description="Year filter e.g. 2026"),
    reg_month: str = Query("", description="Month filter e.g. 1..12"),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)
    if current_user.role == "viewer":
        return RedirectResponse(url="/reports", status_code=302)

    query = db.query(BirthCertificate)

    # Scoping by user role
    if current_user.role == "village_chief" and current_user.village_id:
        query = query.filter(BirthCertificate.village_id == current_user.village_id)
        selected_village_id = current_user.village_id
    else:
        if village_id and village_id.isdigit():
            query = query.filter(BirthCertificate.village_id == int(village_id))
            selected_village_id = int(village_id)
        else:
            selected_village_id = None

    # Text Search Filter
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.filter(
            or_(
                BirthCertificate.name_kh.ilike(search_term),
                BirthCertificate.name_en.ilike(search_term),
                BirthCertificate.certificate_no.ilike(search_term),
                BirthCertificate.book_no.ilike(search_term),
                BirthCertificate.father_name.ilike(search_term),
                BirthCertificate.mother_name.ilike(search_term)
            )
        )

    # Registration Status Filter
    if reg_status == "registered":
        query = query.filter(BirthCertificate.is_registered_voter == True)
    elif reg_status == "unregistered":
        query = query.filter(BirthCertificate.is_registered_voter == False)

    all_records = query.order_by(desc(BirthCertificate.id)).all()

    now = get_cambodia_now()
    current_year = now.year
    current_month = now.month
    current_month_name = f"ខែ{KHMER_MONTH_NAMES.get(current_month, '')}"

    # Filter in Python for eligibility and period categories
    filtered_records = []
    for rec in all_records:
        rec_dt = parse_created_dt(rec)
        
        # Period Filter
        if period == "this_month":
            if not (rec_dt and rec_dt.year == current_year and rec_dt.month == current_month):
                continue
        elif period == "last_month":
            last_m = current_month - 1 if current_month > 1 else 12
            last_m_year = current_year if current_month > 1 else current_year - 1
            if not (rec_dt and rec_dt.year == last_m_year and rec_dt.month == last_m):
                continue
        elif period == "this_year":
            if not (rec_dt and rec_dt.year == current_year):
                continue
        elif period == "last_year":
            if not (rec_dt and rec_dt.year == current_year - 1):
                continue

        # Custom Month / Year Filter
        if reg_year and reg_year.isdigit():
            if not (rec_dt and rec_dt.year == int(reg_year)):
                continue
        if reg_month and reg_month.isdigit():
            if not (rec_dt and rec_dt.month == int(reg_month)):
                continue

        # Eligibility filter
        if eligibility == "turning_18_this_year":
            if rec.is_turning_18_this_year:
                filtered_records.append(rec)
        elif eligibility == "turning_18_next_year":
            if rec.is_turning_18_next_year:
                filtered_records.append(rec)
        elif eligibility == "eligible_now":
            if rec.is_eligible_now:
                filtered_records.append(rec)
        elif eligibility == "under_18":
            if not rec.is_eligible_now and not rec.is_turning_18_this_year:
                filtered_records.append(rec)
        else:
            filtered_records.append(rec)

    total_count = len(filtered_records)
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages

    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    paged_records = filtered_records[start_idx:end_idx]

    # Global KPI Calculations (based on all birth records accessible)
    base_all_query = db.query(BirthCertificate)
    if current_user.role == "village_chief" and current_user.village_id:
        base_all_query = base_all_query.filter(BirthCertificate.village_id == current_user.village_id)
    all_scope_records = base_all_query.all()

    kpi_total = len(all_scope_records)
    
    # Monthly and Yearly KPI
    this_month_recs = [r for r in all_scope_records if parse_created_dt(r) and parse_created_dt(r).year == current_year and parse_created_dt(r).month == current_month]
    kpi_this_month = len(this_month_recs)
    kpi_this_month_female = len([r for r in this_month_recs if r.gender == "ស្រី"])
    kpi_this_month_male = len([r for r in this_month_recs if r.gender == "ប្រុស"])

    this_year_recs = [r for r in all_scope_records if parse_created_dt(r) and parse_created_dt(r).year == current_year]
    kpi_this_year = len(this_year_recs)
    kpi_this_year_female = len([r for r in this_year_recs if r.gender == "ស្រី"])
    kpi_this_year_male = len([r for r in this_year_recs if r.gender == "ប្រុស"])

    kpi_turning_18_this_year = len([r for r in all_scope_records if r.is_turning_18_this_year])
    kpi_eligible_now = len([r for r in all_scope_records if r.is_eligible_now or r.is_turning_18_this_year])
    kpi_unregistered_eligible = len([r for r in all_scope_records if (r.is_eligible_now or r.is_turning_18_this_year) and not r.is_registered_voter])
    kpi_registered_eligible = len([r for r in all_scope_records if (r.is_eligible_now or r.is_turning_18_this_year) and r.is_registered_voter])
    kpi_reg_rate = round((kpi_registered_eligible / kpi_eligible_now * 100), 1) if kpi_eligible_now > 0 else 0.0

    # 12-Month breakdown summary for the selected year (default current_year)
    view_year = int(reg_year) if reg_year and reg_year.isdigit() else current_year
    monthly_stats_summary = []
    for m_idx in range(1, 13):
        m_recs = [r for r in all_scope_records if parse_created_dt(r) and parse_created_dt(r).year == view_year and parse_created_dt(r).month == m_idx]
        monthly_stats_summary.append({
            "month_num": m_idx,
            "month_name": f"ខែ{KHMER_MONTH_NAMES.get(m_idx, '')}",
            "total": len(m_recs),
            "female": len([r for r in m_recs if r.gender == "ស្រី"]),
            "male": len([r for r in m_recs if r.gender == "ប្រុស"]),
            "is_current": (m_idx == current_month and view_year == current_year)
        })

    # Available years for dropdown
    recorded_years = sorted(list(set([parse_created_dt(b).year for b in all_scope_records if parse_created_dt(b)])), reverse=True)
    if current_year not in recorded_years:
        recorded_years.insert(0, current_year)

    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    return templates.TemplateResponse(
        request=request,
        name="birth_certificates/index.html",
        context={
            "current_user": current_user,
            "records": paged_records,
            "total_count": total_count,
            "page": page,
            "total_pages": total_pages,
            "page_size": page_size,
            "q": q,
            "village_id": village_id,
            "selected_village_id": selected_village_id,
            "eligibility": eligibility,
            "reg_status": reg_status,
            "period": period,
            "reg_year": reg_year,
            "reg_month": reg_month,
            "view_year": view_year,
            "available_years": recorded_years,
            "monthly_stats_summary": monthly_stats_summary,
            "villages": villages,
            "stations": stations,
            "kpi_total": kpi_total,
            "kpi_this_month": kpi_this_month,
            "kpi_this_month_female": kpi_this_month_female,
            "kpi_this_month_male": kpi_this_month_male,
            "kpi_this_year": kpi_this_year,
            "kpi_this_year_female": kpi_this_year_female,
            "kpi_this_year_male": kpi_this_year_male,
            "kpi_turning_18_this_year": kpi_turning_18_this_year,
            "kpi_eligible_now": kpi_eligible_now,
            "kpi_unregistered_eligible": kpi_unregistered_eligible,
            "kpi_registered_eligible": kpi_registered_eligible,
            "kpi_reg_rate": kpi_reg_rate,
            "current_year": current_year,
            "current_month": current_month,
            "current_month_name": current_month_name,
            "today_str": get_cambodia_today_str(),
            "khmer_months": KHMER_MONTH_NAMES
        }
    )

@router.get("/birth-certificates/youth-pipeline", response_class=HTMLResponse)
def youth_pipeline_dashboard(
    request: Request,
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()
    all_births = db.query(BirthCertificate).all()

    current_year = get_cambodia_now().year

    total_births = len(all_births)
    turning_18_this_year = [b for b in all_births if b.is_turning_18_this_year]
    turning_18_next_year = [b for b in all_births if b.is_turning_18_next_year]
    eligible_18_plus = [b for b in all_births if b.is_eligible_now]
    under_17 = [b for b in all_births if not b.is_eligible_now and not b.is_turning_18_this_year and not b.is_turning_18_next_year]

    all_eligible = [b for b in all_births if b.is_eligible_now or b.is_turning_18_this_year]
    registered_youth = [b for b in all_eligible if b.is_registered_voter]
    unregistered_youth = [b for b in all_eligible if not b.is_registered_voter]

    reg_rate = round((len(registered_youth) / len(all_eligible) * 100), 1) if all_eligible else 0.0

    # Village by village breakdown
    village_stats = []
    for v in villages:
        v_births = [b for b in all_births if b.village_id == v.id]
        v_eligible = [b for b in v_births if b.is_eligible_now or b.is_turning_18_this_year]
        v_registered = [b for b in v_eligible if b.is_registered_voter]
        v_unregistered = [b for b in v_eligible if not b.is_registered_voter]
        v_turning_next_year = [b for b in v_births if b.is_turning_18_next_year]
        v_rate = round((len(v_registered) / len(v_eligible) * 100), 1) if v_eligible else 0.0
        
        village_stats.append({
            "village": v,
            "total_births": len(v_births),
            "eligible_count": len(v_eligible),
            "registered_count": len(v_registered),
            "unregistered_count": len(v_unregistered),
            "turning_next_year": len(v_turning_next_year),
            "rate": v_rate
        })

    # Sort villages by unregistered count descending for priority outreach
    village_stats.sort(key=lambda x: x["unregistered_count"], reverse=True)

    return templates.TemplateResponse(
        request=request,
        name="birth_certificates/pipeline.html",
        context={
            "current_user": current_user,
            "villages": villages,
            "stations": stations,
            "village_stats": village_stats,
            "total_births": total_births,
            "turning_18_this_year_count": len(turning_18_this_year),
            "turning_18_next_year_count": len(turning_18_next_year),
            "eligible_18_plus_count": len(eligible_18_plus),
            "under_17_count": len(under_17),
            "all_eligible_count": len(all_eligible),
            "registered_youth_count": len(registered_youth),
            "unregistered_youth_count": len(unregistered_youth),
            "reg_rate": reg_rate,
            "current_year": current_year,
            "unregistered_youth_sample": unregistered_youth[:15]
        }
    )

@router.get("/api/birth-certificates/check-duplicate")
def check_duplicate_birth_certificate(
    certificate_no: str = Query("", description="Certificate Number"),
    book_no: str = Query("", description="Book Number"),
    exclude_id: int = Query(0, description="Exclude ID for edit mode"),
    db: Session = Depends(get_db)
):
    cert_clean = certificate_no.strip()
    book_clean = book_no.strip() if book_no else ""

    if not cert_clean:
        return {"duplicate": False, "message": ""}

    query = db.query(BirthCertificate)
    if exclude_id > 0:
        query = query.filter(BirthCertificate.id != exclude_id)

    # 1. Match both Certificate No and Book No if Book No provided
    existing = None
    if book_clean:
        existing = query.filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower(),
            func.lower(func.trim(BirthCertificate.book_no)) == book_clean.lower()
        ).first()

    # 2. Match Certificate No if no exact (cert + book) match or if book_no not provided
    if not existing:
        existing = query.filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower()
        ).first()

    if existing:
        v_name = f"ភូមិ{existing.village.name_kh}" if existing.village else ""
        b_info = f"សៀវភៅលេខ {existing.book_no}" if existing.book_no else ""
        details = " • ".join(filter(None, [f"លេខសំបុត្រ {existing.certificate_no}", b_info, v_name]))
        
        book_mention = f" និងលេខសៀវភៅ '{existing.book_no}'" if existing.book_no else ""
        return {
            "duplicate": True,
            "message": f"លេខសំបុត្រកំណើត '{existing.certificate_no}'{book_mention} ត្រូវបានបញ្ចូលសម្រាប់ឈ្មោះ '{existing.name_kh}' ({details}) រួចហើយ!",
            "existing": {
                "id": existing.id,
                "certificate_no": existing.certificate_no,
                "book_no": existing.book_no or "",
                "name_kh": existing.name_kh,
                "name_en": existing.name_en,
                "dob": existing.dob,
                "village_name": existing.village.name_kh if existing.village else ""
            }
        }

    return {
        "duplicate": False,
        "message": "លេខសំបុត្រកំណើត និងលេខសៀវភៅត្រឹមត្រូវ (មិនស្ទួនក្នុងប្រព័ន្ធឡើយ អាចកត់ត្រាបាន)"
    }

@router.post("/birth-certificates/create")
async def create_birth_certificate(
    request: Request,
    certificate_no: str = Form(...),
    book_no: str = Form(""),
    name_kh: str = Form(...),
    name_en: str = Form(...),
    gender: str = Form(...),
    dob: str = Form(...),
    registered_date: str = Form(""),
    pob: str = Form(""),
    father_name: str = Form(""),
    mother_name: str = Form(""),
    address: str = Form(""),
    village_id: int = Form(...),
    notes: str = Form(""),
    attachment: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    cert_clean = certificate_no.strip()
    book_clean = book_no.strip() if book_no else ""

    # Check duplicate certificate_no and book_no
    existing = None
    if book_clean:
        existing = db.query(BirthCertificate).filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower(),
            func.lower(func.trim(BirthCertificate.book_no)) == book_clean.lower()
        ).first()

    if not existing:
        existing = db.query(BirthCertificate).filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower()
        ).first()

    if existing:
        v_name = f"ភូមិ{existing.village.name_kh}" if existing.village else ""
        book_text = f" និងលេខសៀវភៅ '{existing.book_no}'" if existing.book_no else ""
        return RedirectResponse(
            url=f"/birth-certificates?error=ទិន្នន័យស្ទួន៖ លេខសំបុត្រកំណើត '{existing.certificate_no}'{book_text} ត្រូវបានបញ្ចូលសម្រាប់ឈ្មោះ '{existing.name_kh}' ({v_name}) រួចហើយ!",
            status_code=302
        )

    # Check if there is an existing voter with identical name and dob to auto-link
    matched_voter = db.query(Voter).filter(
        Voter.name_kh == name_kh.strip(),
        Voter.dob == dob.strip(),
        Voter.village_id == village_id
    ).first()

    attachment_url = save_birth_attachment(attachment) if attachment and attachment.filename else None

    reg_date_clean = registered_date.strip() if registered_date and registered_date.strip() else get_cambodia_today_str()

    birth = BirthCertificate(
        certificate_no=cert_clean,
        book_no=book_clean if book_clean else None,
        name_kh=name_kh.strip(),
        name_en=name_en.strip().upper(),
        gender=gender.strip(),
        dob=dob.strip(),
        registered_date=reg_date_clean,
        pob=pob.strip() if pob else None,
        father_name=father_name.strip() if father_name else None,
        mother_name=mother_name.strip() if mother_name else None,
        address=address.strip() if address else None,
        village_id=village_id,
        attachment_url=attachment_url,
        notes=notes.strip() if notes else None,
        is_registered_voter=True if matched_voter else False,
        voter_id=matched_voter.id if matched_voter else None
    )
    db.add(birth)
    db.commit()
    db.refresh(birth)

    log_activity(
        db=db,
        user=current_user,
        action="CREATE_BIRTH_CERT",
        description=f"បានកត់ត្រាសំបុត្រកំណើតលេខ {birth.certificate_no} ឈ្មោះ {birth.name_kh}" + (" (មានឯកសារភ្ជាប់)" if attachment_url else ""),
        target_type="birth_certificate",
        target_id=str(birth.id),
        action_type="success",
        request=request
    )

    return RedirectResponse(
        url="/birth-certificates?msg=បានកត់ត្រាសំបុត្រកំណើតដោយជោគជ័យ",
        status_code=302
    )

@router.post("/birth-certificates/edit/{id}")
async def edit_birth_certificate(
    id: int,
    request: Request,
    certificate_no: str = Form(...),
    book_no: str = Form(""),
    name_kh: str = Form(...),
    name_en: str = Form(...),
    gender: str = Form(...),
    dob: str = Form(...),
    registered_date: str = Form(""),
    pob: str = Form(""),
    father_name: str = Form(""),
    mother_name: str = Form(""),
    address: str = Form(""),
    village_id: int = Form(...),
    notes: str = Form(""),
    attachment: UploadFile = File(None),
    remove_attachment: str = Form("0"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    birth = db.query(BirthCertificate).filter(BirthCertificate.id == id).first()
    if not birth:
        raise HTTPException(status_code=404, detail="រកមិនឃើញសំបុត្រកំណើត")

    cert_clean = certificate_no.strip()
    book_clean = book_no.strip() if book_no else ""

    dup = None
    if book_clean:
        dup = db.query(BirthCertificate).filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower(),
            func.lower(func.trim(BirthCertificate.book_no)) == book_clean.lower(),
            BirthCertificate.id != id
        ).first()

    if not dup:
        dup = db.query(BirthCertificate).filter(
            func.lower(func.trim(BirthCertificate.certificate_no)) == cert_clean.lower(),
            BirthCertificate.id != id
        ).first()

    if dup:
        v_name = f"ភូមិ{dup.village.name_kh}" if dup.village else ""
        book_text = f" និងលេខសៀវភៅ '{dup.book_no}'" if dup.book_no else ""
        return RedirectResponse(
            url=f"/birth-certificates?error=ទិន្នន័យស្ទួន៖ លេខសំបុត្រកំណើត '{dup.certificate_no}'{book_text} មានរួចហើយសម្រាប់ឈ្មោះ '{dup.name_kh}' ({v_name})!",
            status_code=302
        )

    birth.certificate_no = cert_clean
    birth.book_no = book_clean if book_clean else None
    birth.name_kh = name_kh.strip()
    birth.name_en = name_en.strip().upper()
    birth.gender = gender.strip()
    birth.dob = dob.strip()
    if registered_date and registered_date.strip():
        birth.registered_date = registered_date.strip()
    elif not birth.registered_date:
        birth.registered_date = get_cambodia_today_str()
    birth.pob = pob.strip() if pob else None
    birth.father_name = father_name.strip() if father_name else None
    birth.mother_name = mother_name.strip() if mother_name else None
    birth.address = address.strip() if address else None
    birth.village_id = village_id
    birth.notes = notes.strip() if notes else None

    # Handle attachment updates
    if remove_attachment == "1":
        birth.attachment_url = None
    elif attachment and attachment.filename:
        new_url = save_birth_attachment(attachment)
        if new_url:
            birth.attachment_url = new_url

    db.commit()

    log_activity(
        db=db,
        user=current_user,
        action="UPDATE_BIRTH_CERT",
        description=f"បានកែប្រែសំបុត្រកំណើតលេខ {birth.certificate_no} ឈ្មោះ {birth.name_kh}",
        target_type="birth_certificate",
        target_id=str(birth.id),
        action_type="warning",
        request=request
    )

    return RedirectResponse(
        url="/birth-certificates?msg=បានកែប្រែសំបុត្រកំណើតដោយជោគជ័យ",
        status_code=302
    )

@router.post("/birth-certificates/delete/{id}")
async def delete_birth_certificate(
    id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        return RedirectResponse(url="/birth-certificates?error=មានតែ Admin ទើបមានសិទ្ធិលុប", status_code=302)

    birth = db.query(BirthCertificate).filter(BirthCertificate.id == id).first()
    if not birth:
        return RedirectResponse(url="/birth-certificates?error=រកមិនឃើញសំបុត្រកំណើត", status_code=302)

    cert_info = f"{birth.certificate_no} ({birth.name_kh})"
    db.delete(birth)
    db.commit()

    log_activity(
        db=db,
        user=current_user,
        action="DELETE_BIRTH_CERT",
        description=f"បានលុបសំបុត្រកំណើតលេខ {cert_info}",
        target_type="birth_certificate",
        target_id=str(id),
        action_type="danger",
        request=request
    )

    return RedirectResponse(
        url="/birth-certificates?msg=បានលុបសំបុត្រកំណើតដោយជោគជ័យ",
        status_code=302
    )

@router.post("/birth-certificates/convert-to-voter/{id}")
async def convert_to_voter(
    id: int,
    request: Request,
    national_id: str = Form(...),
    station_id: int = Form(...),
    address: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    birth = db.query(BirthCertificate).filter(BirthCertificate.id == id).first()
    if not birth:
        return RedirectResponse(url="/birth-certificates?error=រកមិនឃើញសំបុត្រកំណើត", status_code=302)

    # Sanitize & Validate National ID
    try:
        clean_nat_id = sanitize_id_backend(national_id)
    except HTTPException as e:
        return RedirectResponse(url=f"/birth-certificates?error={e.detail}", status_code=302)

    # Check duplicate national_id in Voters
    dup_voter = db.query(Voter).filter(Voter.national_id == clean_nat_id).first()
    if dup_voter:
        return RedirectResponse(
            url=f"/birth-certificates?error=លេខអត្តសញ្ញាណប័ណ្ណ {clean_nat_id} មានក្នុងបញ្ជីបោះឆ្នោតរួចហើយ (ឈ្មោះ {dup_voter.name_kh})",
            status_code=302
        )

    # Check station exists
    station = db.query(PollingStation).filter(PollingStation.id == station_id).first()
    if not station:
        return RedirectResponse(url="/birth-certificates?error=មិនមានការិយាល័យបោះឆ្នោតនេះទេ", status_code=302)

    # Calculate next list_no and voter_code
    max_list_no = db.query(func.max(Voter.list_no)).filter(Voter.station_id == station_id).scalar() or 0
    next_list_no = max_list_no + 1
    voter_code = f"NP-{station.code}-{next_list_no:04d}"

    # Profile photo avatar fallback
    photo_url = f"/static/images/avatars/female_1.jpg" if birth.gender == "ស្រី" else f"/static/images/avatars/male_1.jpg"

    voter_notes = f"ចុះឈ្មោះដោយស្វ័យប្រវត្តិតាមសំបុត្រកំណើតលេខ {birth.certificate_no}"
    if notes and notes.strip():
        voter_notes += f" | {notes.strip()}"

    voter = Voter(
        voter_code=voter_code,
        list_no=next_list_no,
        national_id=clean_nat_id,
        name_kh=birth.name_kh,
        name_en=birth.name_en,
        gender=birth.gender,
        dob=birth.dob,
        address=address.strip() if address else birth.address,
        village_id=birth.village_id,
        station_id=station_id,
        status="active",
        reg_type="new",
        reg_year=get_cambodia_now().year,
        reg_reason="first_time_18",
        photo_url=photo_url,
        has_voted=False,
        notes=voter_notes,
        created_at=get_cambodia_now()
    )
    db.add(voter)
    db.flush()

    # Link to birth certificate
    birth.is_registered_voter = True
    birth.voter_id = voter.id
    db.commit()

    log_activity(
        db=db,
        user=current_user,
        action="CONVERT_TO_VOTER",
        description=f"បានបំប្លែងសំបុត្រកំណើត {birth.name_kh} ទៅជាអ្នកបោះឆ្នោតកូដ {voter.voter_code} (ការិយាល័យ {station.code})",
        target_type="voter",
        target_id=str(voter.id),
        action_type="success",
        request=request
    )

    return RedirectResponse(
        url=f"/birth-certificates?msg=បានបំប្លែង {birth.name_kh} ទៅជាអ្នកបោះឆ្នោតដោយជោគជ័យ (កូដ {voter.voter_code})",
        status_code=302
    )

@router.get("/birth-certificates/export-excel")
def export_birth_certificates_excel(
    request: Request,
    village_id: str = Query(""),
    eligibility: str = Query("all"),
    reg_status: str = Query("all"),
    period: str = Query("all"),
    reg_year: str = Query(""),
    reg_month: str = Query(""),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    query = db.query(BirthCertificate)
    if current_user.role == "village_chief" and current_user.village_id:
        query = query.filter(BirthCertificate.village_id == current_user.village_id)
    elif village_id and village_id.isdigit():
        query = query.filter(BirthCertificate.village_id == int(village_id))

    if reg_status == "registered":
        query = query.filter(BirthCertificate.is_registered_voter == True)
    elif reg_status == "unregistered":
        query = query.filter(BirthCertificate.is_registered_voter == False)

    all_records = query.order_by(asc(BirthCertificate.village_id), asc(BirthCertificate.dob)).all()

    now = get_cambodia_now()
    current_year = now.year
    current_month = now.month

    filtered = []
    for rec in all_records:
        rec_dt = parse_created_dt(rec)
        
        # Period Filter
        if period == "this_month":
            if not (rec_dt and rec_dt.year == current_year and rec_dt.month == current_month):
                continue
        elif period == "last_month":
            last_m = current_month - 1 if current_month > 1 else 12
            last_m_year = current_year if current_month > 1 else current_year - 1
            if not (rec_dt and rec_dt.year == last_m_year and rec_dt.month == last_m):
                continue
        elif period == "this_year":
            if not (rec_dt and rec_dt.year == current_year):
                continue
        elif period == "last_year":
            if not (rec_dt and rec_dt.year == current_year - 1):
                continue

        # Custom Month / Year Filter
        if reg_year and reg_year.isdigit():
            if not (rec_dt and rec_dt.year == int(reg_year)):
                continue
        if reg_month and reg_month.isdigit():
            if not (rec_dt and rec_dt.month == int(reg_month)):
                continue

        if eligibility == "turning_18_this_year" and not rec.is_turning_18_this_year:
            continue
        if eligibility == "turning_18_next_year" and not rec.is_turning_18_next_year:
            continue
        if eligibility == "eligible_now" and not rec.is_eligible_now:
            continue
        if eligibility == "under_18" and (rec.is_eligible_now or rec.is_turning_18_this_year):
            continue
        filtered.append(rec)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីសំបុត្រកំណើត & យុវជន"

    # Header fonts
    title_font = Font(name="Khmer OS Muol Light", size=14, bold=True, color="0f2b5c")
    subtitle_font = Font(name="Khmer OS Siemreap", size=10, italic=True)
    header_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="Khmer OS Siemreap", size=9)
    bold_cell_font = Font(name="Khmer OS Siemreap", size=9, bold=True)

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    even_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # Title Block
    ws.merge_cells("A1:K1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា • ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A1"].font = Font(name="Khmer OS Muol Light", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = "រដ្ឋបាលឃុំនគរភាស ស្រុកអង្គរជុំ ខេត្តសៀមរាប"
    ws["A2"].font = title_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:K3")
    ws["A3"] = f"បញ្ជីកត់ត្រាសំបុត្រកំណើត និងតាមដានយុវជនគ្រប់អាយុបោះឆ្នោត (ទាញយកថ្ងៃទី {get_cambodia_today()})"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.append([]) # Row 4 empty

    # Table Headers
    headers = [
        "ល.រ", "លេខសំបុត្រកំណើត", "កាលបរិច្ឆេទចុះបញ្ជី", "ឈ្មោះខ្មែរ", "ឈ្មោះឡាតាំង", 
        "ភេទ", "ថ្ងៃខែឆ្នាំកំណើត", "អាយុ", "ស្ថានភាពសិទ្ធិ", 
        "ភូមិ", "ឈ្មោះឪពុក-ម្តាយ", "ស្ថានភាពបោះឆ្នោត"
    ]
    ws.append(headers)
    header_row_idx = 5

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[header_row_idx].height = 28

    for i, b in enumerate(filtered, start=1):
        row_idx = header_row_idx + i
        parents = f"{b.father_name or ''} / {b.mother_name or ''}".strip(" /")
        status_text = f"✅ បានចុះឈ្មោះ ({b.voter.voter_code})" if (b.is_registered_voter and b.voter) else "⚠️ មិនទាន់ចុះឈ្មោះ"
        reg_date_str = b.registered_date_effective or ""
        
        row_data = [
            i,
            b.certificate_no,
            reg_date_str,
            b.name_kh,
            b.name_en,
            b.gender,
            b.dob,
            f"{b.age} ឆ្នាំ",
            b.eligibility_badge["text"],
            b.village.name_kh if b.village else "",
            parents,
            status_text
        ]
        ws.append(row_data)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = bold_cell_font if col_idx in [4, 9, 12] else cell_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_row_fill
            if col_idx in [1, 3, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set Column Widths
    col_widths = [6, 18, 20, 22, 8, 14, 10, 20, 16, 24, 24]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Birth_Records_Youth_NokorPheas_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/birth-certificates/print-eligible", response_class=HTMLResponse)
def print_eligible_action_list(
    request: Request,
    village_id: str = Query(""),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    query = db.query(BirthCertificate).filter(
        BirthCertificate.is_registered_voter == False
    )

    if current_user.role == "village_chief" and current_user.village_id:
        query = query.filter(BirthCertificate.village_id == current_user.village_id)
        selected_village = db.query(Village).filter(Village.id == current_user.village_id).first()
    elif village_id and village_id.isdigit():
        query = query.filter(BirthCertificate.village_id == int(village_id))
        selected_village = db.query(Village).filter(Village.id == int(village_id)).first()
    else:
        selected_village = None

    all_unregistered = query.order_by(asc(BirthCertificate.village_id), asc(BirthCertificate.dob)).all()

    # Keep only those turning 18 this year or eligible 18+
    eligible_unregistered = [
        b for b in all_unregistered 
        if b.is_eligible_now or b.is_turning_18_this_year
    ]

    villages = db.query(Village).order_by(Village.code).all()

    return templates.TemplateResponse(
        request=request,
        name="birth_certificates/print_eligible.html",
        context={
            "current_user": current_user,
            "records": eligible_unregistered,
            "selected_village": selected_village,
            "villages": villages,
            "print_date": get_cambodia_today(),
            "current_year": get_cambodia_now().year
        }
    )

@router.get("/verify/birth/{identifier}", response_class=HTMLResponse)
def public_verify_birth_certificate(
    identifier: str,
    request: Request,
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    ident = identifier.strip()

    record = None
    if ident.isdigit():
        record = db.query(BirthCertificate).filter(BirthCertificate.id == int(ident)).first()
    
    if not record:
        record = db.query(BirthCertificate).filter(
            or_(
                BirthCertificate.certificate_no.ilike(ident),
                BirthCertificate.certificate_no == ident
            )
        ).first()

    voter = None
    if record and record.is_registered_voter and record.voter_id:
        voter = db.query(Voter).filter(Voter.id == record.voter_id).first()

    return templates.TemplateResponse(
        request=request,
        name="birth_certificates/verify.html",
        context={
            "current_user": current_user,
            "record": record,
            "voter": voter,
            "identifier": identifier
        }
    )

@router.get("/api/birth-certificates/lookup-qr")
def api_lookup_birth_qr(
    code: str = Query("", description="QR Code or Certificate number or ID"),
    db: Session = Depends(get_db)
):
    clean_code = (code or "").strip()
    if not clean_code:
        return JSONResponse({"found": False, "message": "សូមបញ្ចូលកូដ ឬស្កេន QR សំបុត្រកំណើត"})

    # Check if full URL was scanned (e.g. https://domain/verify/birth/45)
    m = re.search(r"/verify/birth/([^/?#]+)", clean_code)
    if m:
        clean_code = m.group(1).strip()

    record = None
    if clean_code.isdigit():
        record = db.query(BirthCertificate).filter(BirthCertificate.id == int(clean_code)).first()

    if not record:
        record = db.query(BirthCertificate).filter(
            or_(
                BirthCertificate.certificate_no.ilike(clean_code),
                BirthCertificate.certificate_no == clean_code
            )
        ).first()

    if not record:
        return JSONResponse({"found": False, "message": f"រកមិនឃើញសំបុត្រកំណើតដែលមានកូដ '{clean_code}' ឡើយ"})

    voter_info = None
    if record.is_registered_voter and record.voter:
        voter_info = {
            "id": record.voter.id,
            "voter_code": record.voter.voter_code,
            "station_code": record.voter.station.code if record.voter.station else "",
            "station_name": record.voter.station.name if record.voter.station else "",
            "has_voted": record.voter.has_voted
        }

    return JSONResponse({
        "found": True,
        "record": {
            "id": record.id,
            "certificate_no": record.certificate_no,
            "book_no": record.book_no or "",
            "year": record.registered_year or "",
            "name_kh": record.name_kh,
            "name_en": record.name_en or "",
            "gender": record.gender,
            "dob": record.dob,
            "age": record.age,
            "pob": record.pob or "",
            "father_name": record.father_name or "",
            "mother_name": record.mother_name or "",
            "address": record.address or "",
            "village_id": record.village_id,
            "village_name": record.village.name_kh if record.village else "",
            "village_code": record.village.code if record.village else "",
            "is_registered_voter": record.is_registered_voter,
            "voter": voter_info,
            "is_eligible_now": record.is_eligible_now,
            "eligibility_badge": record.eligibility_badge,
            "registered_date": record.registered_date_effective,
            "attachment_url": record.attachment_url,
            "photo_url": record.photo_display,
            "verify_url": f"/verify/birth/{record.id}"
        }
    })

@router.get("/birth-certificates/{id}/print-card", response_class=HTMLResponse)
def print_birth_certificate_card(
    id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    record = db.query(BirthCertificate).filter(BirthCertificate.id == id).first()
    if not record:
        raise HTTPException(status_code=404, detail="រកមិនឃើញសំបុត្រកំណើតឡើយ")

    voter = None
    if record.is_registered_voter and record.voter_id:
        voter = db.query(Voter).filter(Voter.id == record.voter_id).first()

    return templates.TemplateResponse(
        request=request,
        name="birth_certificates/card.html",
        context={
            "current_user": current_user,
            "record": record,
            "voter": voter,
            "print_date": get_cambodia_today()
        }
    )
