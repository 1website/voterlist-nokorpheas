import os
import shutil
import json
import datetime
from fastapi import APIRouter, Request, Depends, HTTPException, Query, Form, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc, func
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

from app.database import get_db, DB_PATH, BASE_DIR, IS_SQLITE, DB_TYPE_NAME
from app.models import User, Village, PollingStation, Voter, AuditLog
from app.auth import get_current_user_optional, require_admin, require_admin_or_officer
from app.audit import log_activity
from app.timezone_utils import get_cambodia_now, get_cambodia_today, get_cambodia_today_str

router = APIRouter()

templates_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
templates = Jinja2Templates(directory=templates_path)

BACKUPS_DIR = os.path.join(BASE_DIR, "backups")
os.makedirs(BACKUPS_DIR, exist_ok=True)

# -----------------------------------------------------------------------------
# AUDIT LOGS ROUTES
# -----------------------------------------------------------------------------

@router.get("/audit-logs", response_class=HTMLResponse)
def audit_logs_page(
    request: Request,
    q: str = Query("", alias="q"),
    action_filter: str = Query("", alias="action"),
    role_filter: str = Query("", alias="role"),
    date_filter: str = Query("", alias="date"),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=5, le=100),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែអ្នកគ្រប់គ្រងប្រព័ន្ធ (Admin) ប៉ុណ្ណោះដែលអាចមើលកំណត់ត្រាសកម្មភាពបាន")

    query = db.query(AuditLog)

    # Search keyword
    if q and q.strip():
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                AuditLog.description.ilike(search),
                AuditLog.username.ilike(search),
                AuditLog.user_full_name.ilike(search),
                AuditLog.target_id.ilike(search),
                AuditLog.ip_address.ilike(search)
            )
        )

    # Filter action
    if action_filter and action_filter.strip():
        query = query.filter(AuditLog.action == action_filter.strip())

    # Filter role
    if role_filter and role_filter.strip():
        query = query.filter(AuditLog.user_role == role_filter.strip())

    # Filter date
    if date_filter and date_filter.strip():
        clean_d = date_filter.strip()
        try:
            datetime.date.fromisoformat(clean_d)
            query = query.filter(func.date(AuditLog.created_at) == clean_d)
        except ValueError:
            pass

    total_count = query.count()
    total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1

    logs = query.order_by(AuditLog.created_at.desc()).offset((page - 1) * limit).limit(limit).all()

    # Summary Statistics
    today_str = get_cambodia_today_str()
    total_today = db.query(AuditLog).filter(func.date(AuditLog.created_at) == today_str).count()
    total_creates = db.query(AuditLog).filter(AuditLog.action == "CREATE_VOTER").count()
    total_edits = db.query(AuditLog).filter(AuditLog.action.in_(["UPDATE_VOTER", "DELETE_VOTER"])).count()
    total_checkins = db.query(AuditLog).filter(AuditLog.action.in_(["CHECKIN", "UNCHECKIN"])).count()

    # Distinct actions for dropdown filter
    distinct_actions = [r[0] for r in db.query(AuditLog.action).distinct().all() if r[0]]

    return templates.TemplateResponse(request=request, name="system/audit_logs.html", context={
        "current_user": current_user,
        "logs": logs,
        "total_count": total_count,
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "q": q,
        "action_filter": action_filter,
        "role_filter": role_filter,
        "date_filter": date_filter,
        "total_today": total_today,
        "total_creates": total_creates,
        "total_edits": total_edits,
        "total_checkins": total_checkins,
        "distinct_actions": distinct_actions
    })

@router.get("/audit-logs/export/excel")
def export_audit_logs_excel(
    request: Request,
    q: str = Query(""),
    action_filter: str = Query("", alias="action"),
    role_filter: str = Query("", alias="role"),
    date_filter: str = Query("", alias="date"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="គ្មានសិទ្ធិទាញយកទិន្នន័យ (Permission denied)")

    query = db.query(AuditLog)
    if q and q.strip():
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                AuditLog.description.ilike(search),
                AuditLog.username.ilike(search),
                AuditLog.user_full_name.ilike(search),
                AuditLog.target_id.ilike(search)
            )
        )
    if action_filter and action_filter.strip():
        query = query.filter(AuditLog.action == action_filter.strip())
    if role_filter and role_filter.strip():
        query = query.filter(AuditLog.user_role == role_filter.strip())
    if date_filter and date_filter.strip():
        clean_d = date_filter.strip()
        try:
            datetime.date.fromisoformat(clean_d)
            query = query.filter(func.date(AuditLog.created_at) == clean_d)
        except ValueError:
            pass

    logs = query.order_by(AuditLog.created_at.desc()).limit(1000).all()

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit_Logs"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    title_font = Font(name="Hanuman", size=14, bold=True, color="1E3A8A")
    header_font = Font(name="Hanuman", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    regular_font = Font(name="Hanuman", size=9)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title Banner
    ws.merge_cells("A1:G1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា • ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A1"].font = Font(name="Hanuman", size=12, bold=True, color="0F172A")
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:G2")
    ws["A2"] = f"របាយការណ៍កំណត់ត្រាសកម្មភាពមន្ត្រី និងប្រព័ន្ធ (Audit Logs) • ឃុំនគរភាស • កាលបរិច្ឆេទ {get_cambodia_today().strftime('%d/%m/%Y')}"
    ws["A2"].font = title_font
    ws["A2"].alignment = center_align

    # Table Headers
    headers = ["ល.រ", "កាលបរិច្ឆេទ & ម៉ោង", "គណនី (User)", "តួនាទី (Role)", "ប្រភេទសកម្មភាព", "បរិយាយសកម្មភាពលម្អិត", "IP Address"]
    ws.append([]) # Row 3 empty
    ws.append(headers) # Row 4 headers

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 26

    # Rows
    for idx, log in enumerate(logs, 1):
        created_str = log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
        row = [
            idx,
            created_str,
            f"{log.user_full_name} ({log.username})" if log.user_full_name else (log.username or ""),
            log.user_role or "",
            log.action or "",
            log.description or "",
            log.ip_address or ""
        ]
        ws.append(row)
        curr_row = 4 + idx
        ws.row_dimensions[curr_row].height = 22
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 2, 4, 5, 7] else left_align

    # Auto Column Widths
    col_widths = {1: 8, 2: 20, 3: 25, 4: 15, 5: 18, 6: 45, 7: 15}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"audit_logs_{get_cambodia_today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/api/audit-logs/clear")
def clear_audit_logs(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចសម្អាត Logs បាន")

    count = db.query(AuditLog).count()
    db.query(AuditLog).delete()
    db.commit()

    log_activity(db, current_user, "CLEAR_LOGS", f"បានសម្អាតកំណត់ត្រាសកម្មភាពចំនួន {count} ជួរ", "system", action_type="warning", request=request)
    return JSONResponse({"success": True, "message": f"បានសម្អាតកំណត់ត្រាចំនួន {count} ដោយជោគជ័យ"})

# -----------------------------------------------------------------------------
# BACKUP & RESTORE ROUTES
# -----------------------------------------------------------------------------

def restore_from_json_dict(payload: dict, db: Session) -> dict:
    villages_data = payload.get("villages", [])
    stations_data = payload.get("stations", [])
    voters_data = payload.get("voters", [])

    imported_voters = 0
    updated_voters = 0

    # 1. Villages
    village_map = {}
    for v_data in villages_data:
        v_code = v_data.get("code")
        if not v_code:
            continue
        v_obj = db.query(Village).filter(Village.code == v_code).first()
        if not v_obj:
            v_obj = Village(
                code=v_code,
                name_kh=v_data.get("name_kh", ""),
                name_en=v_data.get("name_en", ""),
                chief_name=v_data.get("chief_name"),
                chief_phone=v_data.get("chief_phone"),
                total_households=v_data.get("total_households", 0)
            )
            db.add(v_obj)
            db.flush()
        village_map[v_data.get("id")] = v_obj.id
        village_map[v_code] = v_obj.id

    # 2. Polling Stations
    station_map = {}
    for s_data in stations_data:
        s_code = s_data.get("code")
        if not s_code:
            continue
        s_obj = db.query(PollingStation).filter(PollingStation.code == s_code).first()
        v_id = village_map.get(s_data.get("village_id"))
        if not s_obj:
            s_obj = PollingStation(
                code=s_code,
                name=s_data.get("name", ""),
                location=s_data.get("location", ""),
                capacity=s_data.get("capacity", 600),
                village_id=v_id,
                officer_name=s_data.get("officer_name"),
                officer_phone=s_data.get("officer_phone")
            )
            db.add(s_obj)
            db.flush()
        station_map[s_data.get("id")] = s_obj.id
        station_map[s_code] = s_obj.id

    # 3. Voters
    for vtr in voters_data:
        v_code = vtr.get("voter_code")
        nid = vtr.get("national_id")
        if not v_code or not nid:
            continue
        
        target_v_id = village_map.get(vtr.get("village_id"))
        target_s_id = station_map.get(vtr.get("station_id"))
        if not target_v_id:
            first_v = db.query(Village).first()
            target_v_id = first_v.id if first_v else 1
        if not target_s_id:
            first_s = db.query(PollingStation).first()
            target_s_id = first_s.id if first_s else 1

        existing_voter = db.query(Voter).filter(
            or_(Voter.voter_code == v_code, Voter.national_id == nid)
        ).first()

        if existing_voter:
            existing_voter.list_no = vtr.get("list_no", existing_voter.list_no)
            existing_voter.name_kh = vtr.get("name_kh", existing_voter.name_kh)
            existing_voter.name_en = vtr.get("name_en", existing_voter.name_en)
            existing_voter.gender = vtr.get("gender", existing_voter.gender)
            existing_voter.dob = vtr.get("dob", existing_voter.dob)
            existing_voter.address = vtr.get("address", existing_voter.address)
            if target_v_id:
                existing_voter.village_id = target_v_id
            if target_s_id:
                existing_voter.station_id = target_s_id
            existing_voter.status = vtr.get("status", "active")
            existing_voter.has_voted = vtr.get("has_voted", False)
            existing_voter.photo_url = vtr.get("photo_url", existing_voter.photo_url)
            existing_voter.notes = vtr.get("notes", existing_voter.notes)
            updated_voters += 1
        else:
            new_voter = Voter(
                voter_code=v_code,
                list_no=vtr.get("list_no", 1),
                national_id=nid,
                name_kh=vtr.get("name_kh", ""),
                name_en=vtr.get("name_en", ""),
                gender=vtr.get("gender", "ប្រុស"),
                dob=vtr.get("dob", "1990-01-01"),
                address=vtr.get("address", ""),
                village_id=target_v_id,
                station_id=target_s_id,
                status=vtr.get("status", "active"),
                photo_url=vtr.get("photo_url"),
                has_voted=vtr.get("has_voted", False),
                notes=vtr.get("notes", "")
            )
            db.add(new_voter)
            imported_voters += 1

    db.commit()
    return {
        "imported_voters": imported_voters,
        "updated_voters": updated_voters,
        "total_processed": imported_voters + updated_voters
    }

@router.get("/backup", response_class=HTMLResponse)
def backup_page(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចគ្រប់គ្រងការបម្រុងទុកទិន្នន័យបាន")

    # DB Stats
    if IS_SQLITE:
        db_size_bytes = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
        db_size_kb = round(db_size_bytes / 1024, 1)
        db_size_mb = round(db_size_bytes / (1024 * 1024), 2)
    else:
        db_size_kb = "Cloud"
        db_size_mb = "Managed"

    total_voters = db.query(Voter).count()
    total_villages = db.query(Village).count()
    total_stations = db.query(PollingStation).count()
    total_users = db.query(User).count()
    total_logs = db.query(AuditLog).count()

    # List local backup files in backups/
    backup_files = []
    if os.path.exists(BACKUPS_DIR):
        for f in sorted(os.listdir(BACKUPS_DIR), reverse=True):
            if f.endswith(".db") or f.endswith(".json"):
                f_path = os.path.join(BACKUPS_DIR, f)
                stat = os.stat(f_path)
                backup_files.append({
                    "filename": f,
                    "size_kb": round(stat.st_size / 1024, 1),
                    "created_at": datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    "is_db": f.endswith(".db")
                })

    return templates.TemplateResponse(request=request, name="system/backup.html", context={
        "current_user": current_user,
        "is_sqlite": IS_SQLITE,
        "db_type_name": DB_TYPE_NAME,
        "db_size_kb": db_size_kb,
        "db_size_mb": db_size_mb,
        "total_voters": total_voters,
        "total_villages": total_villages,
        "total_stations": total_stations,
        "total_users": total_users,
        "total_logs": total_logs,
        "backup_files": backup_files
    })

@router.get("/api/backup/download")
def download_sqlite_backup(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចទាញយក Backup បាន")

    if not IS_SQLITE or not os.path.exists(DB_PATH):
        # Redirect to JSON backup if not SQLite
        return download_json_backup(request, db)

    timestamp = get_cambodia_now().strftime("%Y%m%d_%H%M%S")
    download_filename = f"voter_list_backup_{timestamp}.db"

    log_activity(db, current_user, "BACKUP_DOWNLOAD", f"បានទាញយកឯកសារ SQLite Backup៖ {download_filename}", "system", action_type="info", request=request)

    return FileResponse(
        path=DB_PATH,
        filename=download_filename,
        media_type="application/x-sqlite3"
    )

@router.get("/api/backup/download-json")
def download_json_backup(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចទាញយក Backup បាន")

    # Serialize complete database
    villages = [{
        "id": v.id, "code": v.code, "name_kh": v.name_kh, "name_en": v.name_en,
        "chief_name": v.chief_name, "chief_phone": v.chief_phone, "total_households": v.total_households
    } for v in db.query(Village).all()]

    stations = [{
        "id": s.id, "code": s.code, "name": s.name, "location": s.location,
        "capacity": s.capacity, "village_id": s.village_id, "officer_name": s.officer_name, "officer_phone": s.officer_phone
    } for s in db.query(PollingStation).all()]

    voters = [{
        "id": v.id, "voter_code": v.voter_code, "list_no": v.list_no,
        "national_id": v.national_id, "name_kh": v.name_kh, "name_en": v.name_en,
        "gender": v.gender, "dob": v.dob, "address": v.address,
        "village_id": v.village_id, "station_id": v.station_id,
        "status": v.status, "has_voted": v.has_voted, "photo_url": v.photo_url,
        "notes": v.notes,
        "created_at": v.created_at.strftime("%Y-%m-%d %H:%M:%S") if v.created_at else ""
    } for v in db.query(Voter).all()]

    backup_payload = {
        "metadata": {
            "app_name": "ប្រព័ន្ធគ្រប់គ្រងអ្នកចុះឈ្មោះបោះឆ្នោត រដ្ឋបាលឃុំនគរភាស",
            "exported_at": get_cambodia_now().strftime("%Y-%m-%d %H:%M:%S"),
            "exported_by": current_user.username,
            "total_voters": len(voters),
            "total_villages": len(villages),
            "total_stations": len(stations)
        },
        "villages": villages,
        "stations": stations,
        "voters": voters
    }

    timestamp = get_cambodia_now().strftime("%Y%m%d_%H%M%S")
    download_filename = f"voter_data_backup_{timestamp}.json"
    json_bytes = json.dumps(backup_payload, ensure_ascii=False, indent=2).encode('utf-8')

    log_activity(db, current_user, "BACKUP_JSON", f"បានទាញយកឯកសារ JSON Backup៖ {download_filename} ({len(voters)} នាក់)", "system", action_type="info", request=request)

    return Response(
        content=json_bytes,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={download_filename}"}
    )

@router.post("/api/backup/create-snapshot")
def create_local_snapshot(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចបង្កើត Snapshot បាន")

    timestamp = get_cambodia_now().strftime("%Y%m%d_%H%M%S")

    if IS_SQLITE and os.path.exists(DB_PATH):
        snapshot_name = f"snapshot_{timestamp}.db"
        dest_path = os.path.join(BACKUPS_DIR, snapshot_name)
        shutil.copy2(DB_PATH, dest_path)
        file_size_kb = round(os.path.getsize(dest_path) / 1024, 1)
    else:
        # Create JSON snapshot for PostgreSQL or Cloud storage
        snapshot_name = f"snapshot_{timestamp}.json"
        dest_path = os.path.join(BACKUPS_DIR, snapshot_name)
        villages = [{"id": v.id, "code": v.code, "name_kh": v.name_kh, "name_en": v.name_en} for v in db.query(Village).all()]
        stations = [{"id": s.id, "code": s.code, "name": s.name, "village_id": s.village_id} for s in db.query(PollingStation).all()]
        voters = [{
            "id": v.id, "voter_code": v.voter_code, "list_no": v.list_no, "national_id": v.national_id,
            "name_kh": v.name_kh, "name_en": v.name_en, "gender": v.gender, "dob": v.dob,
            "village_id": v.village_id, "station_id": v.station_id, "status": v.status, "has_voted": v.has_voted
        } for v in db.query(Voter).all()]
        payload = {"metadata": {"exported_at": str(datetime.datetime.now())}, "villages": villages, "stations": stations, "voters": voters}
        with open(dest_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        file_size_kb = round(os.path.getsize(dest_path) / 1024, 1)

    log_activity(db, current_user, "BACKUP_SNAPSHOT", f"បានបង្កើត Snapshot ថ្មី៖ {snapshot_name} ({file_size_kb} KB)", "system", action_type="success", request=request)

    return JSONResponse({
        "success": True,
        "message": f"បានបង្កើតច្បាប់ចម្លងបម្រុងទុក (Snapshot) '{snapshot_name}' ដោយជោគជ័យ!",
        "filename": snapshot_name,
        "size_kb": file_size_kb
    })

@router.post("/api/backup/restore")
async def restore_database(
    request: Request,
    backup_file: UploadFile = File(None),
    snapshot_filename: str = Form(None),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user or current_user.role != "admin":
        raise HTTPException(status_code=403, detail="មានតែ Admin ប៉ុណ្ណោះដែលអាចស្តារទិន្នន័យ (Restore) បាន")

    try:
        if backup_file and backup_file.filename:
            filename_lower = backup_file.filename.lower()
            content = await backup_file.read()
            if len(content) < 10:
                raise HTTPException(status_code=400, detail="ឯកសារ Backup គ្មានទិន្នន័យឡើយ")

            # Case A: JSON file restore (Cross-database: works for SQLite & PostgreSQL)
            if filename_lower.endswith(".json"):
                try:
                    payload = json.loads(content.decode("utf-8"))
                    res = restore_from_json_dict(payload, db)
                    msg = f"បានស្តារទិន្នន័យពី JSON ({res['imported_voters']} នាក់ថ្មី, {res['updated_voters']} នាក់កែប្រែ) ដោយជោគជ័យ!"
                    log_activity(db, current_user, "RESTORE_JSON", msg, "system", action_type="warning", request=request)
                    return JSONResponse({"success": True, "message": msg})
                except Exception as e:
                    raise HTTPException(status_code=400, detail=f"ឯកសារ JSON មិនត្រឹមត្រូវ៖ {str(e)}")

            # Case B: SQLite file restore
            elif filename_lower.endswith((".db", ".sqlite", ".sqlite3")):
                if not IS_SQLITE:
                    raise HTTPException(status_code=400, detail="ប្រព័ន្ធកំពុងដំណើរការលើ Cloud PostgreSQL Database។ សូមប្រើប្រាស់ឯកសារ JSON Backup (.json) សម្រាប់ធ្វើការស្តារទិន្នន័យ។")
                
                safety_timestamp = get_cambodia_now().strftime("%Y%m%d_%H%M%S")
                safety_file = os.path.join(BACKUPS_DIR, f"pre_restore_safety_{safety_timestamp}.db")
                if os.path.exists(DB_PATH):
                    shutil.copy2(DB_PATH, safety_file)

                with open(DB_PATH, "wb") as f:
                    f.write(content)

                log_activity(db, current_user, "RESTORE_DATABASE", f"បានស្តារទិន្នន័យពី SQLite '{backup_file.filename}'", "system", action_type="warning", request=request)
                return JSONResponse({"success": True, "message": f"បានស្តារ SQLite Database ពី '{backup_file.filename}' ដោយជោគជ័យ!"})
            else:
                raise HTTPException(status_code=400, detail="សូមជ្រើសរើសឯកសារទម្រង់ .json ឬ .db")

        elif snapshot_filename and snapshot_filename.strip():
            source_name = snapshot_filename.strip()
            source_path = os.path.join(BACKUPS_DIR, source_name)
            if not os.path.exists(source_path):
                raise HTTPException(status_code=404, detail="រកមិនឃើញឯកសារ Snapshot នេះឡើយ")

            if source_name.endswith(".json"):
                with open(source_path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                res = restore_from_json_dict(payload, db)
                msg = f"បានស្តារទិន្នន័យពី Snapshot '{source_name}' ({res['total_processed']} កំណត់ត្រា) ដោយជោគជ័យ!"
                log_activity(db, current_user, "RESTORE_SNAPSHOT", msg, "system", action_type="warning", request=request)
                return JSONResponse({"success": True, "message": msg})
            elif source_name.endswith(".db"):
                if not IS_SQLITE:
                    raise HTTPException(status_code=400, detail="ប្រព័ន្ធកំពុងដំណើរការលើ PostgreSQL។ សូមជ្រើសរើស JSON Snapshot ជំនួសវិញ។")
                shutil.copy2(source_path, DB_PATH)
                log_activity(db, current_user, "RESTORE_SNAPSHOT", f"បានស្តារទិន្នន័យពី SQLite Snapshot '{source_name}'", "system", action_type="warning", request=request)
                return JSONResponse({"success": True, "message": f"បានស្តារទិន្នន័យពី Snapshot '{source_name}' ដោយជោគជ័យ!"})
        else:
            raise HTTPException(status_code=400, detail="សូមជ្រើសរើសឯកសារ Backup (.json ឬ .db) ឬ Snapshot")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"កំហុសក្នុងការស្តារទិន្នន័យ៖ {str(e)}")
