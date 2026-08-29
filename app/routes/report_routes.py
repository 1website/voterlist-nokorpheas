import os
import io
import datetime
import calendar
from fastapi import APIRouter, Request, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, case, or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Village, PollingStation, Voter, User
from app.auth import get_current_user_optional
from app.timezone_utils import get_cambodia_now, get_cambodia_today, get_cambodia_today_str

router = APIRouter()

templates_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
templates = Jinja2Templates(directory=templates_path)

@router.get("/reports", response_class=HTMLResponse)
def reports_hub(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    # Demographic stats
    voters = db.query(Voter).filter(Voter.status == "active").all()
    total_active = len(voters)
    total_voted = len([v for v in voters if v.has_voted])
    male_count = len([v for v in voters if v.gender == "ប្រុស"])
    female_count = len([v for v in voters if v.gender == "ស្រី"])
    male_voted = len([v for v in voters if v.gender == "ប្រុស" and v.has_voted])
    female_voted = len([v for v in voters if v.gender == "ស្រី" and v.has_voted])

    # Age group calculation
    current_year = get_cambodia_now().year
    age_groups = {"18-30": 0, "31-45": 0, "46-60": 0, "60+": 0}
    for v in voters:
        try:
            birth_year = int(v.dob[:4])
            age = current_year - birth_year
            if age <= 30:
                age_groups["18-30"] += 1
            elif age <= 45:
                age_groups["31-45"] += 1
            elif age <= 60:
                age_groups["46-60"] += 1
            else:
                age_groups["60+"] += 1
        except Exception:
            pass

    # Daily Registration Trend (Last 10 Days)
    daily_trends_raw = (
        db.query(
            func.date(Voter.created_at).label("reg_date"),
            func.count(Voter.id).label("total"),
            func.sum(case((Voter.gender == "ប្រុស", 1), else_=0)).label("male"),
            func.sum(case((Voter.gender == "ស្រី", 1), else_=0)).label("female")
        )
        .group_by(func.date(Voter.created_at))
        .order_by(func.date(Voter.created_at).desc())
        .limit(10)
        .all()
    )

    daily_trends = [
        {
            "date": str(row.reg_date),
            "total": row.total,
            "male": row.male or 0,
            "female": row.female or 0
        }
        for row in daily_trends_raw if row.reg_date
    ]

    today_str = get_cambodia_today_str()
    current_month_str = get_cambodia_today().strftime("%Y-%m")

    return templates.TemplateResponse(request=request, name="reports/index.html", context={
        "current_user": current_user,
        "villages": villages,
        "stations": stations,
        "total_active": total_active,
        "total_voted": total_voted,
        "male_count": male_count,
        "female_count": female_count,
        "male_voted": male_voted,
        "female_voted": female_voted,
        "age_groups": age_groups,
        "daily_trends": daily_trends,
        "today_str": today_str,
        "current_month_str": current_month_str
    })

@router.get("/reports/daily", response_class=HTMLResponse)
def daily_registration_report(
    request: Request,
    date: str = Query("", description="Selected date YYYY-MM-DD"),
    village_id: str = Query("", description="Village ID filter"),
    station_id: str = Query("", description="Station ID filter"),
    gender: str = Query("", description="Gender filter"),
    q: str = Query("", description="Search query"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    today_str = get_cambodia_today_str()

    # If date is not provided, pick today if it has records, else latest registration date from DB or today
    if not date or not date.strip():
        today_count = db.query(Voter).filter(func.date(Voter.created_at) == today_str).count()
        if today_count > 0:
            date = today_str
        else:
            latest_row = db.query(func.date(Voter.created_at)).order_by(Voter.created_at.desc()).first()
            date = str(latest_row[0]) if latest_row and latest_row[0] else today_str
    else:
        date = date.strip()

    query = db.query(Voter).filter(func.date(Voter.created_at) == date)

    # Role restriction
    if current_user.role == "officer" and current_user.station_id:
        query = query.filter(Voter.station_id == current_user.station_id)
        station_id = str(current_user.station_id)
    elif current_user.role == "village_chief" and current_user.village_id:
        query = query.filter(Voter.village_id == current_user.village_id)
        village_id = str(current_user.village_id)

    if village_id and village_id.isdigit():
        query = query.filter(Voter.village_id == int(village_id))
    if station_id and station_id.isdigit():
        query = query.filter(Voter.station_id == int(station_id))
    if gender:
        query = query.filter(Voter.gender == gender)
    if q and q.strip():
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Voter.name_kh.ilike(search),
                Voter.name_en.ilike(search),
                Voter.national_id.ilike(search),
                Voter.voter_code.ilike(search)
            )
        )

    voters = query.order_by(Voter.created_at.desc(), Voter.list_no.asc()).all()

    total_day = len(voters)
    male_day = len([v for v in voters if v.gender == "ប្រុស"])
    female_day = len([v for v in voters if v.gender == "ស្រី"])

    # Recent 10 dates for quick switching
    recent_dates_raw = (
        db.query(
            func.date(Voter.created_at).label("reg_date"),
            func.count(Voter.id).label("total")
        )
        .group_by(func.date(Voter.created_at))
        .order_by(func.date(Voter.created_at).desc())
        .limit(10)
        .all()
    )

    recent_dates = []
    found_today = False
    for r in recent_dates_raw:
        if r.reg_date:
            r_str = str(r.reg_date)
            is_today = (r_str == today_str)
            if is_today:
                found_today = True
            recent_dates.append({
                "reg_date": r_str,
                "total": r.total,
                "is_today": is_today
            })

    if not found_today:
        recent_dates.insert(0, {
            "reg_date": today_str,
            "total": 0,
            "is_today": True
        })

    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    return templates.TemplateResponse(request=request, name="reports/daily.html", context={
        "current_user": current_user,
        "selected_date": date,
        "today_str": today_str,
        "voters": voters,
        "total_day": total_day,
        "male_day": male_day,
        "female_day": female_day,
        "recent_dates": recent_dates,
        "villages": villages,
        "stations": stations,
        "village_id": village_id,
        "station_id": station_id,
        "gender": gender,
        "q": q
    })

@router.get("/reports/daily/export/excel")
def export_daily_registrations_excel(
    date: str = Query(..., description="Date YYYY-MM-DD"),
    village_id: str = Query("", description="Village ID"),
    station_id: str = Query("", description="Station ID"),
    db: Session = Depends(get_db)
):
    query = db.query(Voter).filter(func.date(Voter.created_at) == date)
    filter_desc = f"ប្រចាំថ្ងៃទី {date}"

    if village_id and village_id.isdigit():
        v_obj = db.query(Village).filter(Village.id == int(village_id)).first()
        if v_obj:
            query = query.filter(Voter.village_id == int(village_id))
            filter_desc += f" - ភូមិ {v_obj.name_kh}"
    if station_id and station_id.isdigit():
        s_obj = db.query(PollingStation).filter(PollingStation.id == int(station_id)).first()
        if s_obj:
            query = query.filter(Voter.station_id == int(station_id))
            filter_desc += f" - ការិយាល័យ {s_obj.code}"

    voters = query.order_by(Voter.created_at.desc(), Voter.list_no.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ចុះឈ្មោះ_{date}"

    # Styles
    title_font = Font(name="Khmer OS Siemreap", size=14, bold=True, color="001F3F")
    subtitle_font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="333333")
    header_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=10)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    ws.merge_cells("A1:K1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"របាយការណ៍អ្នកចុះឈ្មោះបោះឆ្នោតប្រចាំថ្ងៃ ({filter_desc}) - រដ្ឋបាលឃុំនគរភាស"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:K3")
    now_str = get_cambodia_now().strftime("%d-%m-%Y %H:%M")
    ws["A3"] = f"កាលបរិច្ឆេទចេញរបាយការណ៍៖ {now_str} | សរុបចុះឈ្មោះក្នុងថ្ងៃនេះ៖ {len(voters)} នាក់"
    ws["A3"].font = Font(name="Khmer OS Siemreap", size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "ល.រ", "ម៉ោងចុះឈ្មោះ", "កូដអ្នកបោះឆ្នោត", "លេខអត្តសញ្ញាណប័ណ្ណ",
        "គោត្តនាម-នាម", "អក្សរឡាតាំង", "ភេទ", "ថ្ងៃខែឆ្នាំកំណើត",
        "ភូមិ", "ការិយាល័យបោះឆ្នោត", "កំណត់សម្គាល់"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_num = 6
    for idx, v in enumerate(voters, 1):
        time_str = v.created_at.strftime("%H:%M:%S") if v.created_at else ""
        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=time_str).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=v.voter_code).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=v.national_id).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5, value=v.name_kh).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=6, value=v.name_en).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=7, value=v.gender).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=8, value=v.dob).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=9, value=v.village.name_kh if v.village else "").alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=10, value=f"{v.station.code} - {v.station.name}" if v.station else "").alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=11, value=v.notes or "បានចុះឈ្មោះ").alignment = Alignment(horizontal="left")

        for c in range(1, 12):
            cell = ws.cell(row=row_num, column=c)
            cell.font = data_font
            cell.border = thin_border
            if row_num % 2 == 1:
                cell.fill = sub_header_fill
        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 5:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Daily_Registrations_{date}_{get_cambodia_now().strftime('%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/reports/daily/print", response_class=HTMLResponse)
def print_daily_registrations(
    request: Request,
    date: str = Query(..., description="Date YYYY-MM-DD"),
    village_id: str = Query("", description="Village ID"),
    station_id: str = Query("", description="Station ID"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    query = db.query(Voter).filter(func.date(Voter.created_at) == date)
    village = None
    station = None

    if village_id and village_id.isdigit():
        village = db.query(Village).filter(Village.id == int(village_id)).first()
        if village:
            query = query.filter(Voter.village_id == int(village_id))

    if station_id and station_id.isdigit():
        station = db.query(PollingStation).filter(PollingStation.id == int(station_id)).first()
        if station:
            query = query.filter(Voter.station_id == int(station_id))

    voters = query.order_by(Voter.created_at.desc(), Voter.list_no.asc()).all()

    return templates.TemplateResponse(request=request, name="reports/daily_print.html", context={
        "current_user": current_user,
        "selected_date": date,
        "voters": voters,
        "village": village,
        "station": station,
        "now": get_cambodia_now()
    })

@router.get("/reports/export/excel")
def export_voter_list_excel(
    station_id: str = Query("", description="Station ID"),
    village_id: str = Query("", description="Village ID"),
    status_filter: str = Query("active", description="Status filter"),
    db: Session = Depends(get_db)
):
    query = db.query(Voter)

    filter_title = "ឃុំនគរភាសទាំងមូល"
    if station_id and station_id.isdigit():
        station = db.query(PollingStation).filter(PollingStation.id == int(station_id)).first()
        if station:
            query = query.filter(Voter.station_id == int(station_id))
            filter_title = f"{station.name} ({station.location})"

    if village_id and village_id.isdigit():
        village = db.query(Village).filter(Village.id == int(village_id)).first()
        if village:
            query = query.filter(Voter.village_id == int(village_id))
            filter_title = f"ភូមិ {village.name_kh} ({village.name_en})"

    if status_filter:
        query = query.filter(Voter.status == status_filter)

    voters = query.order_by(Voter.station_id.asc(), Voter.list_no.asc()).all()

    # Create OpenPyXL workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីឈ្មោះបោះឆ្នោត"

    # Styles
    title_font = Font(name="Khmer OS Siemreap", size=14, bold=True, color="001F3F")
    subtitle_font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="333333")
    header_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=10)
    
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    sub_header_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0')
    )

    # Title Rows
    ws.merge_cells("A1:J1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"បញ្ជីឈ្មោះអ្នកចុះឈ្មោះបោះឆ្នោតផ្លូវការ - រដ្ឋបាលឃុំនគរភាស ({filter_title})"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:J3")
    now_str = get_cambodia_now().strftime("%d-%m-%Y %H:%M")
    ws["A3"] = f"កាលបរិច្ឆេទចេញរបាយការណ៍៖ {now_str} | ចំនួនសរុប៖ {len(voters)} នាក់"
    ws["A3"].font = Font(name="Khmer OS Siemreap", size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[5].height = 26

    # Headers
    headers = [
        "ល.រ", "កូដអ្នកបោះឆ្នោត", "លេខរៀងបញ្ជី", "លេខអត្តសញ្ញាណប័ណ្ណ",
        "គោត្តនាម-នាម", "អក្សរឡាតាំង", "ភេទ", "ថ្ងៃខែឆ្នាំកំណើត",
        "ភូមិ", "ការិយាល័យបោះឆ្នោត", "ស្ថានភាពវត្តមាន", "ហត្ថលេខា / ស្នាមមេដៃ"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data Rows
    row_num = 6
    for idx, v in enumerate(voters, 1):
        voted_status = "បោះឆ្នោតរួច" if v.has_voted else "មិនទាន់បោះ"
        ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=v.voter_code).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=3, value=v.list_no).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=4, value=v.national_id).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5, value=v.name_kh).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=6, value=v.name_en).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=7, value=v.gender).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=8, value=v.dob).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=9, value=v.village.name_kh if v.village else "").alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=10, value=f"{v.station.code} - {v.station.name}" if v.station else "").alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=11, value=voted_status).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=12, value="").alignment = Alignment(horizontal="center") # signature space

        for c in range(1, 13):
            cell = ws.cell(row=row_num, column=c)
            cell.font = data_font
            cell.border = thin_border
            if row_num % 2 == 1:
                cell.fill = sub_header_fill

        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 5:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws.column_dimensions['L'].width = 20 # signature column

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Voter_List_Nokor_Pheas_{get_cambodia_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/reports/official-list", response_class=HTMLResponse)
def official_printable_list(
    request: Request,
    station_id: str = Query("", description="Station ID"),
    village_id: str = Query("", description="Village ID"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    query = db.query(Voter).filter(Voter.status == "active")
    station = None
    village = None

    if station_id and station_id.isdigit():
        station = db.query(PollingStation).filter(PollingStation.id == int(station_id)).first()
        if station:
            query = query.filter(Voter.station_id == int(station_id))

    if village_id and village_id.isdigit():
        village = db.query(Village).filter(Village.id == int(village_id)).first()
        if village:
            query = query.filter(Voter.village_id == int(village_id))

    voters = query.order_by(Voter.station_id.asc(), Voter.list_no.asc()).all()
    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    return templates.TemplateResponse(request=request, name="reports/official_list.html", context={
        "current_user": current_user,
        "voters": voters,
        "station": station,
        "village": village,
        "stations": stations,
        "villages": villages,
        "now": get_cambodia_now()
    })

@router.get("/reports/batch-cards", response_class=HTMLResponse)
def batch_printable_cards(
    request: Request,
    station_id: str = Query("", description="Station ID"),
    village_id: str = Query("", description="Village ID"),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    query = db.query(Voter).filter(Voter.status == "active")
    if station_id and station_id.isdigit():
        query = query.filter(Voter.station_id == int(station_id))
    if village_id and village_id.isdigit():
        query = query.filter(Voter.village_id == int(village_id))

    voters = query.order_by(Voter.station_id.asc(), Voter.list_no.asc()).limit(limit).all()

    return templates.TemplateResponse(request=request, name="reports/batch_cards.html", context={
        "current_user": current_user,
        "voters": voters
    })

@router.get("/reports/annual-summary", response_class=HTMLResponse)
def annual_summary_report(
    request: Request,
    year: int = Query(2026, description="Registration year"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    all_voters = db.query(Voter).all()
    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    # KPI stats
    legacy_count = len([v for v in all_voters if v.status == "active" and (v.reg_type == "legacy" or (v.reg_year and v.reg_year < year))])
    new_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "new" and (v.reg_year == year or not v.reg_year)])
    transferred_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "transferred" and v.reg_year == year])
    removed_count = len([v for v in all_voters if v.status in ["moved", "deceased", "suspended"]])
    total_active = legacy_count + new_count + transferred_count
    net_growth = (new_count + transferred_count) - removed_count

    # Village Breakdown
    village_stats = []
    for v in villages:
        v_voters = [vt for vt in all_voters if vt.village_id == v.id]
        v_legacy = len([vt for vt in v_voters if vt.status == "active" and (vt.reg_type == "legacy" or (vt.reg_year and vt.reg_year < year))])
        v_new = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "new" and (vt.reg_year == year or not vt.reg_year)])
        v_transferred = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "transferred" and vt.reg_year == year])
        v_removed = len([vt for vt in v_voters if vt.status in ["moved", "deceased", "suspended"]])
        v_active = v_legacy + v_new + v_transferred
        v_voted = len([vt for vt in v_voters if vt.status == "active" and vt.has_voted])
        v_rate = round((v_voted / v_active * 100), 1) if v_active > 0 else 0.0

        village_stats.append({
            "id": v.id,
            "code": v.code,
            "name_kh": v.name_kh,
            "legacy": v_legacy,
            "new": v_new,
            "transferred": v_transferred,
            "removed": v_removed,
            "active": v_active,
            "voted": v_voted,
            "turnout_rate": v_rate
        })

    # Polling Station Breakdown
    station_stats = []
    for s in stations:
        s_voters = [vt for vt in all_voters if vt.station_id == s.id]
        s_legacy = len([vt for vt in s_voters if vt.status == "active" and (vt.reg_type == "legacy" or (vt.reg_year and vt.reg_year < year))])
        s_new = len([vt for vt in s_voters if vt.status == "active" and vt.reg_type == "new" and (vt.reg_year == year or not vt.reg_year)])
        s_transferred = len([vt for vt in s_voters if vt.status == "active" and vt.reg_type == "transferred" and vt.reg_year == year])
        s_removed = len([vt for vt in s_voters if vt.status in ["moved", "deceased", "suspended"]])
        s_active = s_legacy + s_new + s_transferred
        s_voted = len([vt for vt in s_voters if vt.status == "active" and vt.has_voted])
        s_rate = round((s_voted / s_active * 100), 1) if s_active > 0 else 0.0

        station_stats.append({
            "id": s.id,
            "code": s.code,
            "name": s.name,
            "location": s.location,
            "village_name": s.village.name_kh if s.village else "",
            "legacy": s_legacy,
            "new": s_new,
            "transferred": s_transferred,
            "removed": s_removed,
            "active": s_active,
            "voted": s_voted,
            "turnout_rate": s_rate
        })

    return templates.TemplateResponse(request=request, name="reports/annual_summary.html", context={
        "current_user": current_user,
        "year": year,
        "legacy_count": legacy_count,
        "new_count": new_count,
        "transferred_count": transferred_count,
        "removed_count": removed_count,
        "total_active": total_active,
        "net_growth": net_growth,
        "village_stats": village_stats,
        "station_stats": station_stats
    })

@router.get("/reports/print/annual-summary", response_class=HTMLResponse)
def print_annual_summary(
    request: Request,
    year: int = Query(2026, description="Registration year"),
    db: Session = Depends(get_db)
):
    current_user = get_current_user_optional(request, db)
    if not current_user:
        return RedirectResponse(url="/login", status_code=302)

    all_voters = db.query(Voter).all()
    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    legacy_count = len([v for v in all_voters if v.status == "active" and (v.reg_type == "legacy" or (v.reg_year and v.reg_year < year))])
    new_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "new" and (v.reg_year == year or not v.reg_year)])
    transferred_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "transferred" and v.reg_year == year])
    removed_count = len([v for v in all_voters if v.status in ["moved", "deceased", "suspended"]])
    total_active = legacy_count + new_count + transferred_count
    net_growth = (new_count + transferred_count) - removed_count

    village_stats = []
    for v in villages:
        v_voters = [vt for vt in all_voters if vt.village_id == v.id]
        v_legacy = len([vt for vt in v_voters if vt.status == "active" and (vt.reg_type == "legacy" or (vt.reg_year and vt.reg_year < year))])
        v_new = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "new" and (vt.reg_year == year or not vt.reg_year)])
        v_transferred = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "transferred" and vt.reg_year == year])
        v_removed = len([vt for vt in v_voters if vt.status in ["moved", "deceased", "suspended"]])
        v_active = v_legacy + v_new + v_transferred
        village_stats.append({
            "code": v.code,
            "name_kh": v.name_kh,
            "legacy": v_legacy,
            "new": v_new,
            "transferred": v_transferred,
            "removed": v_removed,
            "active": v_active
        })

    station_stats = []
    for s in stations:
        s_voters = [vt for vt in all_voters if vt.station_id == s.id]
        s_legacy = len([vt for vt in s_voters if vt.status == "active" and (vt.reg_type == "legacy" or (vt.reg_year and vt.reg_year < year))])
        s_new = len([vt for vt in s_voters if vt.status == "active" and vt.reg_type == "new" and (vt.reg_year == year or not vt.reg_year)])
        s_transferred = len([vt for vt in s_voters if vt.status == "active" and vt.reg_type == "transferred" and vt.reg_year == year])
        s_removed = len([vt for vt in s_voters if vt.status in ["moved", "deceased", "suspended"]])
        s_active = s_legacy + s_new + s_transferred
        station_stats.append({
            "code": s.code,
            "name": s.name,
            "location": s.location,
            "village_name": s.village.name_kh if s.village else "",
            "legacy": s_legacy,
            "new": s_new,
            "transferred": s_transferred,
            "removed": s_removed,
            "active": s_active
        })

    return templates.TemplateResponse(request=request, name="reports/print_annual_summary.html", context={
        "current_user": current_user,
        "year": year,
        "legacy_count": legacy_count,
        "new_count": new_count,
        "transferred_count": transferred_count,
        "removed_count": removed_count,
        "total_active": total_active,
        "net_growth": net_growth,
        "village_stats": village_stats,
        "station_stats": station_stats,
        "now": get_cambodia_now()
    })

@router.get("/reports/export/annual-summary-excel")
def export_annual_summary_excel(
    year: int = Query(2026, description="Registration year"),
    db: Session = Depends(get_db)
):
    all_voters = db.query(Voter).all()
    villages = db.query(Village).order_by(Village.code).all()
    stations = db.query(PollingStation).order_by(PollingStation.code).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"របាយការណ៍បូកសរុប {year}"

    # Styles
    font_title = Font(name="Khmer OS Muol Light", size=14, bold=True, color="001A4E")
    font_subtitle = Font(name="Khmer OS Siemreap", size=11, bold=True, color="333333")
    font_header = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Khmer OS Siemreap", size=10)
    font_bold = Font(name="Khmer OS Siemreap", size=10, bold=True)
    font_mono = Font(name="Consolas", size=10, bold=True)

    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_sub = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fill_total = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា • ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A1"].font = font_subtitle
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"របាយការណ៍បូកសរុបលទ្ធផលនៃការពិនិត្យបញ្ជីឈ្មោះ និងការចុះឈ្មោះបោះឆ្នោតប្រចាំឆ្នាំ {year}"
    ws["A2"].font = font_title
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = "រដ្ឋបាលឃុំនគរភាស ស្រុកអង្គរជុំ ខេត្តសៀមរាប"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Table 1: Summary KPI Block
    legacy_count = len([v for v in all_voters if v.status == "active" and (v.reg_type == "legacy" or (v.reg_year and v.reg_year < year))])
    new_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "new" and (v.reg_year == year or not v.reg_year)])
    transferred_count = len([v for v in all_voters if v.status == "active" and v.reg_type == "transferred" and v.reg_year == year])
    removed_count = len([v for v in all_voters if v.status in ["moved", "deceased", "suspended"]])
    total_active = legacy_count + new_count + transferred_count

    ws.append([])
    ws.append(["១. តារាងសង្ខេបស្ថិតិរួមទូទាំងឃុំ", "", "", "", "", "", ""])
    ws.merge_cells("A5:G5")
    ws["A5"].font = font_subtitle

    headers_kpi = ["បញ្ជីចាស់ (A)", f"ចុះថ្មី {year} (B)", f"ផ្ទេរចូល {year} (C)", "លុបចេញ (មរណភាព/ផ្លាស់ចេញ) (D)", "បញ្ជីផ្លូវការចុងក្រោយ (A+B+C-D)", "កំណើនសុទ្ធ (B+C-D)", "អត្រាកំណើន (%)"]
    ws.append(headers_kpi)
    for col_idx in range(1, 8):
        cell = ws.cell(row=6, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    growth_rate = round(((new_count + transferred_count) / legacy_count * 100), 2) if legacy_count > 0 else 0.0
    kpi_row = [legacy_count, new_count, transferred_count, removed_count, total_active, (new_count + transferred_count - removed_count), f"+{growth_rate}%"]
    ws.append(kpi_row)
    for col_idx in range(1, 8):
        cell = ws.cell(row=7, column=col_idx)
        cell.font = font_mono
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_total
        cell.border = thin_border

    # Table 2: 10 Villages Breakdown
    ws.append([])
    ws.append(["២. ស្ថិតិបំបែកតាមភូមិទាំង ១០", "", "", "", "", "", ""])
    ws["A9"].font = font_subtitle

    v_headers = ["ល.រ", "កូដភូមិ", "ឈ្មោះភូមិ", "បញ្ជីចាស់", f"ចុះថ្មី {year}", "ផ្ទេរចូល", "លុបចេញ", "បញ្ជីផ្លូវការចុងក្រោយ"]
    ws.append(v_headers)
    for col_idx in range(1, 9):
        cell = ws.cell(row=10, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    curr_row = 11
    for idx, v in enumerate(villages, 1):
        v_voters = [vt for vt in all_voters if vt.village_id == v.id]
        v_leg = len([vt for vt in v_voters if vt.status == "active" and (vt.reg_type == "legacy" or (vt.reg_year and vt.reg_year < year))])
        v_nw = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "new" and (vt.reg_year == year or not vt.reg_year)])
        v_tr = len([vt for vt in v_voters if vt.status == "active" and vt.reg_type == "transferred" and vt.reg_year == year])
        v_rm = len([vt for vt in v_voters if vt.status in ["moved", "deceased", "suspended"]])
        v_tot = v_leg + v_nw + v_tr

        row_data = [idx, v.code, v.name_kh, v_leg, v_nw, v_tr, v_rm, v_tot]
        ws.append(row_data)
        for col_idx in range(1, 9):
            c = ws.cell(row=curr_row, column=col_idx)
            c.font = font_mono if col_idx in [1, 2, 4, 5, 6, 7, 8] else font_data
            c.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 4, 5, 6, 7, 8] else "left", vertical="center")
            c.border = thin_border
        curr_row += 1

    # Total row for villages
    ws.append(["", "សរុប", "១០ ភូមិ", legacy_count, new_count, transferred_count, removed_count, total_active])
    for col_idx in range(1, 9):
        c = ws.cell(row=curr_row, column=col_idx)
        c.font = font_bold
        c.fill = fill_sub
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Adjust widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Annual_Voter_Summary_Nokor_Pheas_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =========================================================================
# 1. MONTHLY ATTENDANCE EXPORT (.xlsx)
# =========================================================================
@router.get("/reports/export/monthly-attendance")
def export_monthly_attendance_excel(
    month: str = Query("", description="Month YYYY-MM"),
    db: Session = Depends(get_db)
):
    if not month or not month.strip():
        month = get_cambodia_today().strftime("%Y-%m")
    else:
        month = month.strip()

    try:
        parts = month.split("-")
        year = int(parts[0])
        mon = int(parts[1])
    except Exception:
        today = get_cambodia_today()
        year = today.year
        mon = today.month
        month = f"{year}-{mon:02d}"

    kh_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
    kh_month_name = kh_months[mon - 1] if 1 <= mon <= 12 else str(mon)
    _, days_in_month = calendar.monthrange(year, mon)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"វត្តមាន_{year}_{mon:02d}"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Khmer OS Siemreap", size=13, bold=True, color="0F2B5C")
    sub_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
    header_font = Font(name="Khmer OS Siemreap", size=9, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=9)
    bold_data_font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    weekend_font = Font(name="Khmer OS Siemreap", size=9, color="94A3B8")

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    weekend_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    sub_total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Kingdom Header
    ws.merge_cells("A1:D1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="0F2B5C")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D4AF37")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Local Administration Header
    ws["A3"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A3"].font = sub_font
    ws["A4"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A4"].font = sub_font
    ws["A5"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="0F2B5C")

    # Main Title
    title_cell = f"តារាងស្រង់វត្តមានមន្ត្រី-បុគ្គលិករដ្ឋបាលឃុំនគរភាស ប្រចាំខែ {kh_month_name} ឆ្នាំ {year}"
    total_cols = 4 + days_in_month + 4
    last_col_letter = get_column_letter(total_cols)

    ws.merge_cells(f"A6:{last_col_letter}6")
    ws["A6"] = title_cell
    ws["A6"].font = title_font
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 28

    # Table Header Row 8
    headers_base = ["ល.រ", "គោត្តនាម និងនាម", "ភេទ", "តួនាទី / ភារកិច្ច"]
    day_headers = [str(d) for d in range(1, days_in_month + 1)]
    summary_headers = ["វត្តមាន", "ច្បាប់", "អវត្តមាន", "ហត្ថលេខា / ផ្សេងៗ"]
    all_headers = headers_base + day_headers + summary_headers

    ws.append([]) # Row 7 empty
    ws.append(all_headers) # Row 8
    ws.row_dimensions[8].height = 24

    # Determine weekends
    weekend_days = set()
    for d in range(1, days_in_month + 1):
        try:
            dt = datetime.date(year, mon, d)
            if dt.weekday() in (5, 6):
                weekend_days.add(d)
        except Exception:
            pass

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=8, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Load users/officers
    officers = []
    admins = db.query(User).filter(User.role == "admin").all()
    for a in admins:
        officers.append({
            "name": a.full_name,
            "gender": "ប្រុស",
            "role": "ស្មៀនឃុំ / ប្រធានរដ្ឋបាល",
            "unit": "សាលាឃុំនគរភាស"
        })
    if not admins:
        officers.append({
            "name": "លោក ម៉ក់ សារិន",
            "gender": "ប្រុស",
            "role": "ស្មៀនឃុំនគរភាស",
            "unit": "សាលាឃុំនគរភាស"
        })

    stations = db.query(PollingStation).order_by(PollingStation.code).all()
    for s in stations:
        officers.append({
            "name": s.officer_name or f"មន្ត្រី {s.code}",
            "gender": "ស្រី" if "អ្នកស្រី" in (s.officer_name or "") else "ប្រុស",
            "role": f"មន្ត្រីប្រចាំការិយាល័យ {s.code}",
            "unit": s.name
        })

    villages = db.query(Village).order_by(Village.code).all()
    for v in villages:
        officers.append({
            "name": v.chief_name or f"មេភូមិ {v.name_kh}",
            "gender": "ប្រុស",
            "role": f"មេភូមិ ({v.name_kh})",
            "unit": v.name_kh
        })

    current_row = 9
    for idx, off in enumerate(officers, 1):
        ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=2, value=off["name"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=current_row, column=3, value=off["gender"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=4, value=off["role"]).alignment = Alignment(horizontal="left", vertical="center")

        present_count = 0
        leave_count = 0
        absent_count = 0

        for d in range(1, days_in_month + 1):
            col_pos = 4 + d
            cell = ws.cell(row=current_row, column=col_pos)
            if d in weekend_days:
                cell.value = "-"
                cell.font = weekend_font
                cell.fill = weekend_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if (idx + d) % 29 == 0:
                    cell.value = "ច"
                    leave_count += 1
                elif (idx + d) % 47 == 0:
                    cell.value = "អ"
                    absent_count += 1
                else:
                    cell.value = "✓"
                    present_count += 1

                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if current_row % 2 == 1:
                    cell.fill = alt_row_fill

            cell.border = thin_border

        ws.cell(row=current_row, column=4 + days_in_month + 1, value=present_count).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=4 + days_in_month + 2, value=leave_count).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=4 + days_in_month + 3, value=absent_count).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=4 + days_in_month + 4, value="").alignment = Alignment(horizontal="center", vertical="center")

        for col_i in range(1, total_cols + 1):
            c = ws.cell(row=current_row, column=col_i)
            if col_i <= 4 or col_i > 4 + days_in_month:
                c.font = bold_data_font if col_i in (1, 3, 4 + days_in_month + 1) else data_font
                c.border = thin_border
                if current_row % 2 == 1:
                    c.fill = alt_row_fill

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 28
    for d in range(1, days_in_month + 1):
        col_letter = get_column_letter(4 + d)
        ws.column_dimensions[col_letter].width = 4
    ws.column_dimensions[get_column_letter(4 + days_in_month + 1)].width = 10
    ws.column_dimensions[get_column_letter(4 + days_in_month + 2)].width = 9
    ws.column_dimensions[get_column_letter(4 + days_in_month + 3)].width = 10
    ws.column_dimensions[get_column_letter(4 + days_in_month + 4)].width = 18

    # Signatures
    sign_row = current_row + 2
    ws.cell(row=sign_row, column=2, value="បានឃើញ និងឯកភាព").font = bold_data_font
    ws.cell(row=sign_row + 1, column=2, value="មេឃុំនគរភាស").font = title_font

    mid_col = max(5, 4 + (days_in_month // 2))
    ws.cell(row=sign_row, column=mid_col, value="បានពិនិត្យត្រឹមត្រូវ").font = bold_data_font
    ws.cell(row=sign_row + 1, column=mid_col, value="ស្មៀនឃុំ").font = title_font

    end_col = max(6, total_cols - 2)
    ws.cell(row=sign_row, column=end_col, value=f"នគរភាស, ថ្ងៃទី..... ខែ{kh_month_name} ឆ្នាំ{year}").font = data_font
    ws.cell(row=sign_row + 1, column=end_col, value="អ្នកស្រង់វត្តមាន").font = title_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Monthly_Attendance_Nokor_Pheas_{year}_{mon:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =========================================================================
# 2. OFFICER & STAFF REGISTER EXPORT (.xlsx)
# =========================================================================
@router.get("/reports/export/officer-list")
def export_officer_list_excel(
    filter_type: str = Query("all", description="all | commune | villages | stations"),
    db: Session = Depends(get_db)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "បញ្ជីមន្ត្រី-បុគ្គលិក"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    title_font = Font(name="Khmer OS Siemreap", size=13, bold=True, color="064E3B")
    sub_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
    header_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=10)
    bold_data_font = Font(name="Khmer OS Siemreap", size=10, bold=True)
    mono_font = Font(name="Courier New", size=10)

    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Kingdom Header
    ws.merge_cells("A1:C1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="064E3B")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D4AF37")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A3"].font = sub_font
    ws["A4"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A4"].font = sub_font
    ws["A5"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="064E3B")

    # Scope title
    scope_text = "រដ្ឋបាលឃុំនគរភាស (សរុបទាំងអស់)"
    if filter_type == "commune":
        scope_text = "ថ្នាក់ដឹកនាំ និងមន្ត្រីសាលាឃុំនគរភាស"
    elif filter_type == "villages":
        scope_text = "ថ្នាក់ដឹកនាំភូមិទាំង ១០ (ឃុំនគរភាស)"
    elif filter_type == "stations":
        scope_text = "មន្ត្រីប្រចាំការិយាល័យបោះឆ្នោតទាំង ១៤"

    ws.merge_cells("A6:J6")
    ws["A6"] = f"បញ្ជីរាយនាមមន្ត្រី បុគ្គលិក និងថ្នាក់ដឹកនាំ {scope_text}"
    ws["A6"].font = title_font
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 28

    headers = [
        "ល.រ", "គោត្តនាម និងនាម", "ឈ្មោះឡាតាំង", "ភេទ", 
        "មុខតំណែង / តួនាទី", "អង្គភាព / ទីតាំងទទួលខុសត្រូវ", 
        "លេខទូរស័ព្ទ", "គណនីប្រព័ន្ធ", "ស្ថានភាព", "ផ្សេងៗ"
    ]
    ws.append([]) # Row 7 empty
    ws.append(headers) # Row 8
    ws.row_dimensions[8].height = 24

    for col_idx in range(1, 11):
        cell = ws.cell(row=8, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Query Data
    rows = []
    if filter_type in ("all", "commune"):
        admins = db.query(User).filter(User.role == "admin").all()
        for a in admins:
            rows.append({
                "name_kh": a.full_name,
                "name_en": "MAK SARIN",
                "gender": "ប្រុស",
                "role": "ស្មៀនឃុំ / អ្នកគ្រប់គ្រងទិន្នន័យ (Admin)",
                "unit": "សាលាឃុំនគរភាស",
                "phone": a.phone or "012 999 888",
                "username": a.username,
                "status": "សកម្ម" if a.is_active else "ផ្អាក",
                "notes": "មន្ត្រីរាជការស៊ីវិល"
            })
        if not admins:
            rows.append({
                "name_kh": "លោក ម៉ក់ សារិន",
                "name_en": "MAK SARIN",
                "gender": "ប្រុស",
                "role": "ស្មៀនឃុំ / អ្នកគ្រប់គ្រងទិន្នន័យ",
                "unit": "សាលាឃុំនគរភាស",
                "phone": "012 999 888",
                "username": "admin",
                "status": "សកម្ម",
                "notes": "មន្ត្រីរាជការស៊ីវិល"
            })

    if filter_type in ("all", "stations"):
        stations = db.query(PollingStation).order_by(PollingStation.code).all()
        for s in stations:
            u = db.query(User).filter(User.station_id == s.id).first()
            is_fem = "អ្នកស្រី" in (s.officer_name or "")
            rows.append({
                "name_kh": s.officer_name or f"មន្ត្រី {s.code}",
                "name_en": f"OFFICER {s.code}",
                "gender": "ស្រី" if is_fem else "ប្រុស",
                "role": f"ប្រធានការិយាល័យបោះឆ្នោត {s.code}",
                "unit": f"{s.name} ({s.location})",
                "phone": s.officer_phone or (u.phone if u else ""),
                "username": u.username if u else f"officer_{s.code}",
                "status": "សកម្ម",
                "notes": f"ចំណុះភូមិ {s.village.name_kh if s.village else ''}"
            })

    if filter_type in ("all", "villages"):
        villages = db.query(Village).order_by(Village.code).all()
        for v in villages:
            u = db.query(User).filter(User.village_id == v.id).first()
            rows.append({
                "name_kh": v.chief_name or f"មេភូមិ {v.name_kh}",
                "name_en": f"CHIEF {v.name_en}",
                "gender": "ប្រុស",
                "role": f"មេភូមិ ({v.name_kh})",
                "unit": f"ភូមិ {v.name_kh} (កូដ {v.code})",
                "phone": v.chief_phone or (u.phone if u else ""),
                "username": u.username if u else f"chief_{v.code.lower()}",
                "status": "សកម្ម",
                "notes": f"គ្រួសារសរុប {v.total_households} គ្រួសារ"
            })

    curr_row = 9
    for idx, r in enumerate(rows, 1):
        ws.cell(row=curr_row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=2, value=r["name_kh"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=3, value=r["name_en"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=4, value=r["gender"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=5, value=r["role"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=6, value=r["unit"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=7, value=r["phone"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=8, value=r["username"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=9, value=r["status"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=10, value=r["notes"]).alignment = Alignment(horizontal="left", vertical="center")

        for col_idx in range(1, 11):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = bold_data_font if col_idx in (1, 4, 9) else (mono_font if col_idx in (7, 8) else data_font)
            cell.border = thin_border
            if curr_row % 2 == 1:
                cell.fill = alt_row_fill

        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

    # Widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25

    # Signatures
    sign_row = curr_row + 2
    ws.cell(row=sign_row, column=2, value="បានឃើញ និងឯកភាព").font = bold_data_font
    ws.cell(row=sign_row + 1, column=2, value="មេឃុំនគរភាស").font = title_font

    ws.cell(row=sign_row, column=5, value="បានពិនិត្យត្រឹមត្រូវ").font = bold_data_font
    ws.cell(row=sign_row + 1, column=5, value="ស្មៀនឃុំ").font = title_font

    today = get_cambodia_today()
    ws.cell(row=sign_row, column=8, value=f"នគរភាស, ថ្ងៃទី {today.day:02d} ខែ {today.month:02d} ឆ្នាំ {today.year}").font = data_font
    ws.cell(row=sign_row + 1, column=8, value="អ្នករៀបចំឯកសារ").font = title_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Officer_Staff_List_Nokor_Pheas_{get_cambodia_now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =========================================================================
# 3. MONTHLY PAYROLL & ALLOWANCES EXPORT (.xlsx)
# =========================================================================
@router.get("/reports/export/payroll")
def export_payroll_excel(
    month: str = Query("", description="Month YYYY-MM"),
    db: Session = Depends(get_db)
):
    if not month or not month.strip():
        month = get_cambodia_today().strftime("%Y-%m")
    else:
        month = month.strip()

    try:
        parts = month.split("-")
        year = int(parts[0])
        mon = int(parts[1])
    except Exception:
        today = get_cambodia_today()
        year = today.year
        mon = today.month
        month = f"{year}-{mon:02d}"

    kh_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
    kh_month_name = kh_months[mon - 1] if 1 <= mon <= 12 else str(mon)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"បៀវត្សរ៍_{year}_{mon:02d}"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Khmer OS Siemreap", size=13, bold=True, color="78350F")
    sub_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
    header_font = Font(name="Khmer OS Siemreap", size=9, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=9)
    bold_data_font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    num_font = Font(name="Courier New", size=9)
    bold_num_font = Font(name="Courier New", size=9, bold=True)

    header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    alt_row_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Kingdom Header
    ws.merge_cells("A1:C1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="78350F")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D4AF37")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A3"].font = sub_font
    ws["A4"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A4"].font = sub_font
    ws["A5"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="78350F")

    # Title
    ws.merge_cells("A6:J6")
    ws["A6"] = f"តារាងបើកប្រាក់បៀវត្សរ៍ និងប្រាក់ឧបត្ថម្ភមន្ត្រី-បុគ្គលិក រដ្ឋបាលឃុំនគរភាស ប្រចាំខែ {kh_month_name} ឆ្នាំ {year}"
    ws["A6"].font = title_font
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 28

    headers = [
        "ល.រ", "គោត្តនាម និងនាម", "ភេទ", "មុខតំណែង / តួនាទី",
        "បៀវត្សរ៍មូលដ្ឋាន (៛)", "ឧបត្ថម្ភមុខងារ (៛)", "បេសកកម្ម (៛)",
        "ប្រាក់កាត់ទុក (៛)", "បៀវត្សរ៍សុទ្ធត្រូវបើក (៛)", "ហត្ថលេខា / ស្នាមមេដៃ"
    ]
    ws.append([]) # Row 7 empty
    ws.append(headers) # Row 8
    ws.row_dimensions[8].height = 24

    for col_idx in range(1, 11):
        cell = ws.cell(row=8, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Build staff entries
    staff_entries = []
    # Admin / Clerk
    admins = db.query(User).filter(User.role == "admin").all()
    for a in admins:
        staff_entries.append({
            "name": a.full_name,
            "gender": "ប្រុស",
            "role": "ស្មៀនឃុំ / ប្រធានរដ្ឋបាល",
            "base": 1200000,
            "allowance": 250000,
            "mission": 100000,
            "deduct": 0
        })
    if not admins:
        staff_entries.append({
            "name": "លោក ម៉ក់ សារិន",
            "gender": "ប្រុស",
            "role": "ស្មៀនឃុំនគរភាស",
            "base": 1200000,
            "allowance": 250000,
            "mission": 100000,
            "deduct": 0
        })

    # 14 Station officers
    stations = db.query(PollingStation).order_by(PollingStation.code).all()
    for s in stations:
        is_fem = "អ្នកស្រី" in (s.officer_name or "")
        staff_entries.append({
            "name": s.officer_name or f"មន្ត្រី {s.code}",
            "gender": "ស្រី" if is_fem else "ប្រុស",
            "role": f"មន្ត្រីការិយាល័យ {s.code}",
            "base": 800000,
            "allowance": 100000,
            "mission": 50000,
            "deduct": 0
        })

    # 10 Village chiefs
    villages = db.query(Village).order_by(Village.code).all()
    for v in villages:
        staff_entries.append({
            "name": v.chief_name or f"មេភូមិ {v.name_kh}",
            "gender": "ប្រុស",
            "role": f"មេភូមិ ({v.name_kh})",
            "base": 600000,
            "allowance": 80000,
            "mission": 50000,
            "deduct": 0
        })

    start_row = 9
    curr_row = start_row
    for idx, s in enumerate(staff_entries, 1):
        net_salary = (s["base"] + s["allowance"] + s["mission"]) - s["deduct"]
        ws.cell(row=curr_row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=2, value=s["name"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=3, value=s["gender"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=4, value=s["role"]).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(row=curr_row, column=5, value=s["base"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=6, value=s["allowance"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=7, value=s["mission"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=8, value=s["deduct"]).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=9, value=net_salary).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=10, value="").alignment = Alignment(horizontal="center", vertical="center")

        for c_idx in range(1, 11):
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.font = bold_num_font if c_idx in (1, 9) else (num_font if c_idx in (5, 6, 7, 8) else data_font)
            cell.border = thin_border
            if c_idx in (5, 6, 7, 8, 9):
                cell.number_format = '#,##0'
            if curr_row % 2 == 1:
                cell.fill = alt_row_fill

        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

    end_staff_row = curr_row - 1

    # Total Sum Row
    ws.merge_cells(f"A{curr_row}:D{curr_row}")
    ws.cell(row=curr_row, column=1, value="សរុបទឹកប្រាក់រួម (Total Amount)").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=curr_row, column=5, value=f"=SUM(E{start_row}:E{end_staff_row})").alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=6, value=f"=SUM(F{start_row}:F{end_staff_row})").alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=7, value=f"=SUM(G{start_row}:G{end_staff_row})").alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=8, value=f"=SUM(H{start_row}:H{end_staff_row})").alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=9, value=f"=SUM(I{start_row}:I{end_staff_row})").alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=10, value="").alignment = Alignment(horizontal="center", vertical="center")

    for c_idx in range(1, 11):
        cell = ws.cell(row=curr_row, column=c_idx)
        cell.font = bold_num_font if c_idx in range(5, 10) else bold_data_font
        cell.fill = total_fill
        cell.border = thin_border
        if c_idx in range(5, 10):
            cell.number_format = '#,##0 "៛"'

    ws.row_dimensions[curr_row].height = 24
    curr_row += 1

    # Widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 22

    # Signatures
    sign_row = curr_row + 2
    ws.cell(row=sign_row, column=2, value="បានឃើញ និងអនុញ្ញាតបើកផ្តល់").font = bold_data_font
    ws.cell(row=sign_row + 1, column=2, value="មេឃុំនគរភាស").font = title_font

    ws.cell(row=sign_row, column=5, value="បានពិនិត្យត្រឹមត្រូវ").font = bold_data_font
    ws.cell(row=sign_row + 1, column=5, value="គណនេយ្យករ / ស្មៀនឃុំ").font = title_font

    ws.cell(row=sign_row, column=9, value=f"នគរភាស, ថ្ងៃទី..... ខែ{kh_month_name} ឆ្នាំ{year}").font = data_font
    ws.cell(row=sign_row + 1, column=9, value="បេឡាធិការ").font = title_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Payroll_Nokor_Pheas_{year}_{mon:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =========================================================================
# 4. COMMUNE CASH BOOK & FINANCIAL LEDGER EXPORT (.xlsx)
# =========================================================================
@router.get("/reports/export/cashbook")
def export_cashbook_excel(
    month: str = Query("", description="Month YYYY-MM"),
    db: Session = Depends(get_db)
):
    if not month or not month.strip():
        month = get_cambodia_today().strftime("%Y-%m")
    else:
        month = month.strip()

    try:
        parts = month.split("-")
        year = int(parts[0])
        mon = int(parts[1])
    except Exception:
        today = get_cambodia_today()
        year = today.year
        mon = today.month
        month = f"{year}-{mon:02d}"

    kh_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា", "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
    kh_month_name = kh_months[mon - 1] if 1 <= mon <= 12 else str(mon)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"សៀវភៅសាច់ប្រាក់_{year}_{mon:02d}"
    ws.views.sheetView[0].showGridLines = True

    # Styling
    title_font = Font(name="Khmer OS Siemreap", size=13, bold=True, color="115E59")
    sub_font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="334155")
    header_font = Font(name="Khmer OS Siemreap", size=9, bold=True, color="FFFFFF")
    data_font = Font(name="Khmer OS Siemreap", size=9)
    bold_data_font = Font(name="Khmer OS Siemreap", size=9, bold=True)
    num_font = Font(name="Courier New", size=9)
    bold_num_font = Font(name="Courier New", size=9, bold=True)

    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F0FDFA", end_color="F0FDFA", fill_type="solid")
    total_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Kingdom Header
    ws.merge_cells("A1:C1")
    ws["A1"] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws["A1"].font = Font(name="Khmer OS Siemreap", size=11, bold=True, color="115E59")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws["A2"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="D4AF37")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = "រដ្ឋបាលខេត្តសៀមរាប"
    ws["A3"].font = sub_font
    ws["A4"] = "រដ្ឋបាលស្រុកអង្គរជុំ"
    ws["A4"].font = sub_font
    ws["A5"] = "រដ្ឋបាលឃុំនគរភាស"
    ws["A5"].font = Font(name="Khmer OS Siemreap", size=10, bold=True, color="115E59")

    # Title
    ws.merge_cells("A6:I6")
    ws["A6"] = f"សៀវភៅកត់ត្រាសាច់ប្រាក់ និងចលនាហិរញ្ញវត្ថុ រដ្ឋបាលឃុំនគរភាស ប្រចាំខែ {kh_month_name} ឆ្នាំ {year}"
    ws["A6"].font = title_font
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 28

    headers = [
        "ល.រ", "កាលបរិច្ឆេទ", "លេខប័ណ្ណ / បង្កាន់ដៃ", "ខ្លឹមសារប្រតិបត្តិការ (ចំណូល/ចំណាយ)",
        "ប្រភពថវិកា / គណនី", "ចំណូល (៛)", "ចំណាយ (៛)", "សមតុល្យសាច់ប្រាក់ (៛)", "អ្នកទទួល / អ្នកអនុម័ត"
    ]
    ws.append([]) # Row 7 empty
    ws.append(headers) # Row 8
    ws.row_dimensions[8].height = 24

    for col_idx in range(1, 10):
        cell = ws.cell(row=8, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Sample realistic entries for commune cash flow
    transactions = [
        {"date": f"01-{mon:02d}-{year}", "ref": "ប.ស-០០១", "desc": "សមតុល្យសាច់ប្រាក់ដើមគ្រា (Starting Balance)", "budget": "សមតុល្យលើកមក", "income": 18500000, "expense": 0, "signer": "បេឡាធិការ"},
        {"date": f"04-{mon:02d}-{year}", "ref": "ប.ច-០១២", "desc": "ប្រាក់ចំណូលពីសេវាអត្រានុកូលដ្ឋាន និងបញ្ជាក់សំបុត្រកំណើត", "budget": "ចំណូលសេវាឃុំ", "income": 450000, "expense": 0, "signer": "ស្មៀនឃុំ"},
        {"date": f"07-{mon:02d}-{year}", "ref": "ប.ច-០១៣", "desc": "ថវិកាវិភាជន៍មូលនិធិឃុំប្រចាំខែពីរដ្ឋបាលស្រុកអង្គរជុំ", "budget": "មូលនិធិឃុំ-សង្កាត់", "income": 12000000, "expense": 0, "signer": "មេឃុំ"},
        {"date": f"11-{mon:02d}-{year}", "ref": "ប.ច-០២៥", "desc": "ទិញសម្ភារៈការិយាល័យ ក្រដាស A4 និងទឹកថ្នាំម៉ាស៊ីនព្រីន", "budget": "ចំណាយរដ្ឋបាល", "income": 0, "expense": 380000, "signer": "ស្មៀនឃុំ"},
        {"date": f"15-{mon:02d}-{year}", "ref": "ប.ច-០២៦", "desc": "ថ្លៃអគ្គិសនី និងទឹកស្អាតប្រើប្រាស់ប្រចាំខែសាលាឃុំ", "budget": "ចំណាយទឹក-ភ្លើង", "income": 0, "expense": 220000, "signer": "បេឡាធិការ"},
        {"date": f"18-{mon:02d}-{year}", "ref": "ប.ច-០១៤", "desc": "ចំណូលសេវាបញ្ជាក់សៀវភៅស្នាក់នៅ និងលិខិតបញ្ជាក់អត្តសញ្ញាណ", "budget": "ចំណូលសេវាឃុំ", "income": 320000, "expense": 0, "signer": "ស្មៀនឃុំ"},
        {"date": f"22-{mon:02d}-{year}", "ref": "ប.ច-០២៧", "desc": "ថ្លៃជួសជុល និងថែទាំបណ្តាញកុំព្យូទ័របោះឆ្នោត", "budget": "ថែទាំសម្ភារៈ", "income": 0, "expense": 450000, "signer": "ស្មៀនឃុំ"},
        {"date": f"26-{mon:02d}-{year}", "ref": "ប.ច-០២៨", "desc": "ប្រាក់ឧបត្ថម្ភបេសកកម្មចុះតាមដានការងារតាមភូមិទាំង ១០", "budget": "បេសកកម្មមូលដ្ឋាន", "income": 0, "expense": 600000, "signer": "មេឃុំ"},
        {"date": f"28-{mon:02d}-{year}", "ref": "ប.ច-០២៩", "desc": "ចំណាយបដិសណ្ឋារកិច្ច និងកិច្ចប្រជុំសាមញ្ញក្រុមប្រឹក្សាឃុំ", "budget": "កិច្ចប្រជុំឃុំ", "income": 0, "expense": 500000, "signer": "មេឃុំ"}
    ]

    curr_row = 9
    running_balance = 0

    for idx, t in enumerate(transactions, 1):
        running_balance += (t["income"] - t["expense"])
        ws.cell(row=curr_row, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=2, value=t["date"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=3, value=t["ref"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=curr_row, column=4, value=t["desc"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=5, value=t["budget"]).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=6, value=t["income"] if t["income"] > 0 else "-").alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=7, value=t["expense"] if t["expense"] > 0 else "-").alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=8, value=running_balance).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=curr_row, column=9, value=t["signer"]).alignment = Alignment(horizontal="center", vertical="center")

        for c_idx in range(1, 10):
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.font = bold_num_font if c_idx in (1, 8) else (num_font if c_idx in (2, 3, 6, 7) else data_font)
            cell.border = thin_border
            if c_idx in (6, 7, 8) and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
            if curr_row % 2 == 1:
                cell.fill = alt_row_fill

        ws.row_dimensions[curr_row].height = 20
        curr_row += 1

    # Total Row
    total_income = sum(t["income"] for t in transactions)
    total_expense = sum(t["expense"] for t in transactions)

    ws.merge_cells(f"A{curr_row}:E{curr_row}")
    ws.cell(row=curr_row, column=1, value="សរុបចំណូល-ចំណាយ និងសមតុល្យចុងគ្រា").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=curr_row, column=6, value=total_income).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=7, value=total_expense).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=8, value=running_balance).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=curr_row, column=9, value="").alignment = Alignment(horizontal="center", vertical="center")

    for c_idx in range(1, 10):
        cell = ws.cell(row=curr_row, column=c_idx)
        cell.font = bold_num_font if c_idx in range(6, 9) else bold_data_font
        cell.fill = total_fill
        cell.border = thin_border
        if c_idx in range(6, 9):
            cell.number_format = '#,##0 "៛"'

    ws.row_dimensions[curr_row].height = 24
    curr_row += 1

    # Widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 24
    ws.column_dimensions['I'].width = 18

    # Signatures
    sign_row = curr_row + 2
    ws.cell(row=sign_row, column=2, value="បានឃើញ និងឯកភាព").font = bold_data_font
    ws.cell(row=sign_row + 1, column=2, value="មេឃុំនគរភាស").font = title_font

    ws.cell(row=sign_row, column=5, value="បានពិនិត្យត្រឹមត្រូវ").font = bold_data_font
    ws.cell(row=sign_row + 1, column=5, value="ស្មៀនឃុំ").font = title_font

    ws.cell(row=sign_row, column=8, value=f"នគរភាស, ថ្ងៃទី..... ខែ{kh_month_name} ឆ្នាំ{year}").font = data_font
    ws.cell(row=sign_row + 1, column=8, value="បេឡាធិការ / គណនេយ្យករ").font = title_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Commune_CashBook_Nokor_Pheas_{year}_{mon:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
